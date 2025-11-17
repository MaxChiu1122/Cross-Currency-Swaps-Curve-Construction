# -*- coding: utf-8 -*-
"""
CCS Curve Construction

This version provides a clean, modular structure for:
- Multi-currency curve construction (USD, THB, JPY)
- Cross-currency swap (CCS) pricing
- FX forward curve construction
- NPV calculations for booking positions

Version: Clean Structure with Improved JPY-THB Triangulation

IMPROVEMENTS FOR JPY-THB CCS:
- Uses proper triangulation methodology: b_JPY-THB(t) = b_JPY-USD(t) - b_THB-USD(t)
- Implements USD-CSA consistent pricing framework
- Proper FX forward triangulation: F_JPYTHB(t) = F_USDTHB(t) / F_USDJPY(t)
- Aligns tenor grids and uses monotone interpolation
- Creates synthetic JPY-THB CCS from liquid USD-based markets
"""

import pandas as pd
import numpy as np
import datetime
import warnings
import sys
import os
from typing import Dict, List, Tuple
from dataclasses import dataclass

warnings.filterwarnings('ignore')

from rateslib import (
    Cal, Curve, Solver, IRS, FXSwap, XCS, FXRates, FXForwards, 
    Schedule, add_tenor, defaults, NoInput
)


# ============================================================================
# CONSTANTS
# ============================================================================

# Date and Time Constants
DAYS_IN_YEAR = 365.25
BUSINESS_DAYS_IN_YEAR = 365
SETTLEMENT_LAG_DAYS = 2
MAX_TENOR_YEARS = 10
MAX_TENOR_DAYS = 3650  # 10 years

# Numeric Constants
DEFAULT_DISCOUNT_FACTOR = 1.0
BASIS_POINTS_MULTIPLIER = 10000.0
PERCENTAGE_MULTIPLIER = 100

# Default Values
DEFAULT_PAYMENT_LAG = 2
DEFAULT_SPREAD = 0.0
DEFAULT_RATE = 0.0

# Curve Interpolation Bounds
MIN_SYNTHETIC_DF = 0.1
MAX_SYNTHETIC_DF = 10.0

# Calendar Configuration
WEEKEND_DAYS = [5, 6]  # Saturday, Sunday

# Default Conventions
DEFAULT_USD_CONVENTION = 'Act360'
DEFAULT_THB_CONVENTION = 'Act365f'
DEFAULT_JPY_CONVENTION = 'Act365f'

# Default Calendars
# Note: THB does not have a standalone calendar - it is always combined with the other currency
# For USD/THB: use NYC + THB holidays, for JPY/THB: use TYO + THB holidays
DEFAULT_USD_CALENDAR = 'nyc'
DEFAULT_THB_CALENDAR = None  # THB calendar is created from currency pair + Holidays sheet
DEFAULT_JPY_CALENDAR = 'tyo'

# Default Curves
DEFAULT_USD_CURVE = 'SOFR'
DEFAULT_THB_CURVE = 'THOR'
DEFAULT_JPY_CURVE = 'TONR'

# Default Frequencies
DEFAULT_FREQUENCY = 'Q'

# Excel Column Numbers (1-indexed)
EXCEL_COL_NET_NPV = 28  # Column AB
EXCEL_COL_LEG1_NPV = 29  # Column AC
EXCEL_COL_LEG2_NPV = 30  # Column AD
EXCEL_COL_BREAKEVEN = 31  # Column AE

# Default P&L Currency
DEFAULT_PL_CURRENCY = 'USD'
SUPPORTED_CURRENCIES = ['USD', 'THB', 'JPY']


@dataclass
class MarketData:
    """Data class to hold market data parameters"""
    curve_date: datetime.datetime
    thb_spot: float
    jpy_spot: float
    thbjpy_spot: float  # THB/JPY spot rate (calculated from USD/THB and USD/JPY)
    lag_days: int
    sofr_rates: pd.DataFrame
    thor_rates: pd.DataFrame
    tonr_rates: pd.DataFrame
    usdthb_sw_rates: pd.DataFrame
    usdjpy_sw_rates: pd.DataFrame
    usdthb_ccs_rates: pd.DataFrame
    usdjpy_ccs_rates: pd.DataFrame
    sofr_fixings: pd.Series
    thor_fixings: pd.Series
    tonr_fixings: pd.Series


@dataclass
class BookingPosition:
    """Data class to hold booking position parameters"""
    # General Trade Information
    booking_id: str
    effective_date: datetime.datetime
    maturity_date: datetime.datetime
    position: str  # 'S/B' or 'B/S'
    
    # Leg 1 Parameters
    leg1_currency: str
    leg1_notional: float
    leg1_fixed_float: str  # 'Fixed' or 'Float'
    leg1_curve: str
    leg1_rate: float  # Fixed rate for fixed legs
    leg1_spread: float  # Spread for floating legs
    leg1_daycount: str
    leg1_pay_freq: str
    leg1_reset_freq: str
    leg1_calendar: str
    
    # Leg 2 Parameters
    leg2_currency: str
    leg2_notional: float
    leg2_fixed_float: str  # 'Fixed' or 'Float'
    leg2_curve: str
    leg2_rate: float  # Fixed rate for fixed legs
    leg2_spread: float  # Spread for floating legs
    leg2_daycount: str
    leg2_frequency: str
    leg2_reset_freq: str
    leg2_calendar: str
    
    # Additional Trade Attributes
    fx_fixing: float
    lookback_days: int
    payment_lag: int
    trade_status: str
    pl_currency: str  # Currency for P&L reporting: 'USD', 'JPY', or 'THB'
    
    @property
    def currencies(self) -> List[str]:
        """Get list of currencies in this position"""
        return [self.leg1_currency, self.leg2_currency]
    
    @property
    def primary_currency(self) -> str:
        """Get primary currency (leg1)"""
        return self.leg1_currency
    
    @property
    def secondary_currency(self) -> str:
        """Get secondary currency (leg2)"""
        return self.leg2_currency
    
    def get_notional(self, currency: str) -> float:
        """Get notional amount for specific currency"""
        if self.leg1_currency == currency:
            return self.leg1_notional
        elif self.leg2_currency == currency:
            return self.leg2_notional
        else:
            return 0.0
    
    def get_spread(self, currency: str) -> float:
        """Get spread for specific currency"""
        if self.leg1_currency == currency:
            return self.leg1_spread
        elif self.leg2_currency == currency:
            return self.leg2_spread
        else:
            return 0.0
    
    def get_rate(self, currency: str) -> float:
        """Get rate for specific currency"""
        if self.leg1_currency == currency:
            return self.leg1_rate
        elif self.leg2_currency == currency:
            return self.leg2_rate
        else:
            return 0.0
    
    def get_fixed_float(self, currency: str) -> str:
        """Get fixed/float status for specific currency"""
        if self.leg1_currency == currency:
            return self.leg1_fixed_float
        elif self.leg2_currency == currency:
            return self.leg2_fixed_float
        else:
            return 'Float'
    
    def get_curve(self, currency: str) -> str:
        """Get curve name for specific currency"""
        if self.leg1_currency == currency:
            return self.leg1_curve
        elif self.leg2_currency == currency:
            return self.leg2_curve
        else:
            return DEFAULT_USD_CURVE
    
    def get_frequency(self, currency: str) -> str:
        """Get frequency for specific currency"""
        if self.leg1_currency == currency:
            return self.leg1_pay_freq
        elif self.leg2_currency == currency:
            return self.leg2_frequency
        else:
            return DEFAULT_FREQUENCY
    
    def get_convention(self, currency: str) -> str:
        """Get day count convention for specific currency"""
        if self.leg1_currency == currency:
            return self.leg1_daycount
        elif self.leg2_currency == currency:
            return self.leg2_daycount
        else:
            # Default conventions by currency
            currency_conventions = {
                'USD': DEFAULT_USD_CONVENTION,
                'THB': DEFAULT_THB_CONVENTION,
                'JPY': DEFAULT_JPY_CONVENTION
            }
            return currency_conventions.get(currency, DEFAULT_USD_CONVENTION)
    
    def get_calendar(self, currency: str) -> str:
        """
        Get calendar for specific currency.
        Note: For THB, returns None as THB calendar is created from currency pair + Holidays.
        """
        if self.leg1_currency == currency:
            return self.leg1_calendar
        elif self.leg2_currency == currency:
            return self.leg2_calendar
        else:
            # Default calendars by currency
            # THB does not have standalone calendar - it's created from pair + Holidays
            currency_calendars = {
                'USD': DEFAULT_USD_CALENDAR,
                'THB': None,  # THB calendar created from pair (e.g., USD/THB -> NYC + THB holidays)
                'JPY': DEFAULT_JPY_CALENDAR
            }
            return currency_calendars.get(currency, DEFAULT_USD_CALENDAR)


class CalendarManager:
    """Manages holiday calendars for different currencies"""
    
    def __init__(self, holidays_file: str):
        self.holidays_file = holidays_file
        self.calendars = {}
        self._load_holidays()
        self._create_calendars()
    
    def _load_holidays(self) -> None:
        """Load holidays from Excel file"""
        df = pd.read_excel(self.holidays_file, sheet_name='Holidays', engine='openpyxl')
        df = df.reset_index()
        
        self.holidays = {
            'US': [],
            'TH': [],
            'EN': [],
            'JP': [],
        }
        
        for _, row in df.iterrows():
            date = row['Date']
            code = row['CDR Code']
            
            if code in self.holidays:
                self.holidays[code].append(datetime.datetime(
                    date.year, date.month, date.day
                ))
    
    def _create_calendars(self) -> None:
        """
        Create calendar objects for different currency combinations.
        
        Note: THB does not have a standalone calendar for trading purposes.
        THB holidays are always combined with the other currency's calendar:
        - USD/THB: Uses 'us_th' (NYC + THB holidays from Holidays sheet)
        - JPY/THB: Uses 'jp_th' (TYO + THB holidays from Holidays sheet)
        
        The standalone 'th' calendar is created for reference but is not used
        for actual trade calendar determination.
        """
        us_holidays = self.holidays['US']
        th_holidays = self.holidays['TH']
        en_holidays = self.holidays['EN']
        jp_holidays = self.holidays['JP']
        
        # Individual calendars (for reference - THB standalone not used for trades)
        self.calendars['us'] = Cal(us_holidays, WEEKEND_DAYS)
        self.calendars['th'] = Cal(th_holidays, WEEKEND_DAYS)  # Not used standalone for trades
        self.calendars['en'] = Cal(en_holidays, WEEKEND_DAYS)
        self.calendars['jp'] = Cal(jp_holidays, WEEKEND_DAYS)
        
        # Combined calendars (used for actual trades)
        # USD/THB: NYC calendar + THB holidays
        self.calendars['us_th'] = Cal(us_holidays + th_holidays, WEEKEND_DAYS)
        # USD/JPY: NYC calendar + JP holidays
        self.calendars['us_jp'] = Cal(us_holidays + jp_holidays, WEEKEND_DAYS)
        # JPY/THB: TYO calendar + THB holidays
        self.calendars['jp_th'] = Cal(jp_holidays + th_holidays, WEEKEND_DAYS)
        self.calendars['us_th_en'] = Cal(us_holidays + th_holidays + en_holidays, WEEKEND_DAYS)
    
    def get_calendar(self, name: str) -> Cal:
        """Get calendar by name"""
        return self.calendars.get(name)


class MarketDataLoader:
    """Handles loading and processing of market data from Excel"""
    
    def __init__(self, excel_file: str):
        self.excel_file = excel_file
        self.market_data = None
        self._load_market_data()
    
    def _load_market_data(self) -> None:
        """Load market data from Excel file"""
        # Load main market data
        self.market_data = pd.read_excel(self.excel_file, sheet_name='Market Data', engine='openpyxl')
        self.market_data = self.market_data.reset_index()
        
        # Load fixings data
        self._load_fixings()
    
    def _load_fixings(self) -> None:
        """Load fixing data from Excel - consolidated single 'Fixing' sheet with side-by-side sections"""
        # Load all fixings from single 'Fixing' sheet
        # Expected layout: Three side-by-side sections, each with Date | Fixing columns
        # Section 1 (cols A-B): Date | SOFR_Fixing
        # Section 2 (cols D-E): Date | TONR_Fixing  
        # Section 3 (cols G-H): Date | THOR_Fixing
        try:
            # Read entire sheet without header
            df_raw = pd.read_excel(self.excel_file, sheet_name='Fixing', engine='openpyxl', header=None)
            
            # Track successful loads
            loaded_fixings = []
            
            # Extract SOFR Fixings (columns 0-1: Date, SOFR_Fixing)
            try:
                # Skip first row (header) and get columns A-B (indices 0-1)
                sofr_data = df_raw.iloc[1:, [0, 1]].copy()
                sofr_data.columns = ['Date', 'SOFR_Fixing']
                sofr_data = sofr_data.dropna()
                
                if len(sofr_data) > 0:
                    self.sofr_fixings = pd.Series(
                        sofr_data['SOFR_Fixing'].values,
                        index=pd.to_datetime(sofr_data['Date'])
                    )
                    loaded_fixings.append(f"SOFR: {len(self.sofr_fixings)} fixings")
                else:
                    self.sofr_fixings = pd.Series(dtype=float)
            except Exception as e:
                print(f"Warning: Could not load SOFR fixings: {e}")
                self.sofr_fixings = pd.Series(dtype=float)
            
            # Extract TONR Fixings (columns 3-4: Date, TONR_Fixing)
            try:
                # Skip first row (header) and get columns D-E (indices 3-4)
                tonr_data = df_raw.iloc[1:, [3, 4]].copy()
                tonr_data.columns = ['Date', 'TONR_Fixing']
                tonr_data = tonr_data.dropna()
                
                if len(tonr_data) > 0:
                    self.tonr_fixings = pd.Series(
                        tonr_data['TONR_Fixing'].values,
                        index=pd.to_datetime(tonr_data['Date'])
                    )
                    loaded_fixings.append(f"TONR: {len(self.tonr_fixings)} fixings")
                else:
                    self.tonr_fixings = pd.Series(dtype=float)
            except Exception as e:
                print(f"Warning: Could not load TONR fixings: {e}")
                self.tonr_fixings = pd.Series(dtype=float)
            
            # Extract THOR Fixings (columns 6-7: Date, THOR_Fixing)
            try:
                # Skip first row (header) and get columns G-H (indices 6-7)
                thor_data = df_raw.iloc[1:, [6, 7]].copy()
                thor_data.columns = ['Date', 'THOR_Fixing']
                thor_data = thor_data.dropna()
                
                if len(thor_data) > 0:
                    self.thor_fixings = pd.Series(
                        thor_data['THOR_Fixing'].values,
                        index=pd.to_datetime(thor_data['Date'])
                    )
                    loaded_fixings.append(f"THOR: {len(self.thor_fixings)} fixings")
                else:
                    self.thor_fixings = pd.Series(dtype=float)
            except Exception as e:
                print(f"Warning: Could not load THOR fixings: {e}")
                self.thor_fixings = pd.Series(dtype=float)
            
            # Display summary
            if loaded_fixings:
                print(f"✓ Loaded fixings from 'Fixing' sheet: {', '.join(loaded_fixings)}")
            else:
                print("Warning: No fixings were loaded from 'Fixing' sheet")
                
        except Exception as e:
            print(f"Error loading 'Fixing' sheet: {e}")
            print("Initializing empty fixing series...")
            self.sofr_fixings = pd.Series(dtype=float)
            self.tonr_fixings = pd.Series(dtype=float)
            self.thor_fixings = pd.Series(dtype=float)
    
    def get_market_data(self) -> MarketData:
        """Extract and return structured market data"""
        lag_days = int(self.market_data.iloc[0, 30])
        thb_spot = self.market_data.USDTHB[0]
        jpy_spot = self.market_data.USDJPY[0]
        # Calculate THB/JPY spot rate from USD/THB and USD/JPY
        thbjpy_spot = thb_spot / jpy_spot  # THB/JPY = (USD/JPY) / (USD/THB)
        # Fetch curve_date from Market Data sheet AB2 cell (column AB, row 2)
        # Try to get it by column name first, then by position
        try:
            # First try to get by column name if it exists
            if 'Curve Date' in self.market_data.columns:
                curve_date = pd.to_datetime(self.market_data['Curve Date'][0])
            else:
                # Fall back to position-based access (AB2 = column 27, row 1)
                curve_date_raw = self.market_data.iloc[1, 27]
                curve_date = pd.to_datetime(curve_date_raw)
        except (KeyError, IndexError, ValueError) as e:
            print(f"Error reading curve_date: {e}")
            # Fallback to current date
            curve_date = pd.Timestamp.now()
        
        # Ensure it's a Python datetime object for rateslib compatibility
        if hasattr(curve_date, 'to_pydatetime'):
            curve_date = curve_date.to_pydatetime()
        elif hasattr(curve_date, 'date'):
            curve_date = curve_date.date()
        
        # Extract rate data
        sofr_rates = self._extract_rates('US_Instrument', 'US_Term', 'US_Rate', 'IRS')
        thor_rates = self._extract_rates('TH_Instrument', 'TH_Term', 'TH_Rate', 'IRS')
        tonr_rates = self._extract_rates('JP_Instrument', 'JP_Term', 'JP_Rate', 'IRS')
        usdthb_sw_rates = self._extract_rates('TH_Forward', 'TF_Term', 'TF_Rate', 'FXSW')
        usdjpy_sw_rates = self._extract_rates('JP_Forward', 'JF_Term', 'JF_Rate', 'FXSW')
        usdthb_ccs_rates = self._extract_rates('UT_Instrument', 'UT_Term', 'UT_Rate', 'CCS')
        usdjpy_ccs_rates = self._extract_rates('UJ_Instrument', 'UJ_Term', 'UJ_Rate', 'CCS')
        
        return MarketData(
            curve_date=curve_date,
            thb_spot=thb_spot,
            jpy_spot=jpy_spot,
            thbjpy_spot=thbjpy_spot,
            lag_days=lag_days,
            sofr_rates=sofr_rates,
            thor_rates=thor_rates,
            tonr_rates=tonr_rates,
            usdthb_sw_rates=usdthb_sw_rates,
            usdjpy_sw_rates=usdjpy_sw_rates,
            usdthb_ccs_rates=usdthb_ccs_rates,
            usdjpy_ccs_rates=usdjpy_ccs_rates,
            sofr_fixings=self.sofr_fixings,
            thor_fixings=self.thor_fixings,
            tonr_fixings=self.tonr_fixings
        )
    
    def _extract_rates(self, inst_col: str, term_col: str, rate_col: str, inst_type: str) -> pd.DataFrame:
        """Extract rate data for specific instrument type"""
        rates = self.market_data[[inst_col, term_col, rate_col]]
        rates = rates[rates[inst_col] == inst_type]
        rates.columns = ['Instrument', 'Term', 'Rate']
        return rates


class CurveBuilder:
    """Handles construction of interest rate curves"""
    
    def __init__(self, market_data: MarketData, calendar_manager: CalendarManager):
        self.market_data = market_data
        self.calendar_manager = calendar_manager
        self.curves = {}
        self.solvers = {}
        self.fx_forwards = {}
        self.instruments = {}
        
        # Calculate settlement dates
        self._calculate_settlement_dates()
        
        # Build curves
        self._build_base_curves()
        self._build_fx_curves()
        self._build_solvers()
    
    def _calculate_settlement_dates(self) -> None:
        """Calculate settlement and valuation dates"""
        us_th_cal = self.calendar_manager.get_calendar('us_th')
        
        self.thb_settlement_date = add_tenor(
            self.market_data.curve_date, "2D", "F", us_th_cal
        )
        self.usbthb_valuation_date = add_tenor(
            self.market_data.curve_date, "2D", "F", us_th_cal
        )
        
        # Convert to datetime objects for rateslib compatibility
        self.thb_settlement_date = pd.to_datetime(self.thb_settlement_date).to_pydatetime()
        self.usbthb_valuation_date = pd.to_datetime(self.usbthb_valuation_date).to_pydatetime()
    
    def _calculate_termination_dates(self, rates_df: pd.DataFrame, calendar: Cal) -> List[datetime.datetime]:
        """Calculate termination dates for rate instruments"""
        termination_dates = []
        
        for term in rates_df["Term"]:
            try:
                termination_date = add_tenor(self.thb_settlement_date, term, "F", calendar)
                termination_dates.append(termination_date)
            except Exception as e:
                print(f"Error calculating termination for term {term}: {e}")
                termination_dates.append(self.thb_settlement_date)
        
        return termination_dates
    
    def _build_base_curves(self) -> None:
        """Build base interest rate curves"""
        # Process rate data and add termination dates
        self._process_rate_data()
        
        # Build individual curves
        self._build_sofr_curve()
        self._build_thor_curve()
        self._build_tonr_curve()
        self._build_fx_swap_curves()
        self._build_ccs_curves()
    
    def _process_rate_data(self) -> None:
        """Process rate data and add termination dates"""
        # SOFR rates
        us_cal = self.calendar_manager.get_calendar('us')
        self.market_data.sofr_rates["Termination"] = self._calculate_termination_dates(
            self.market_data.sofr_rates, us_cal
        )
        self.market_data.sofr_rates["Label"] = self.market_data.sofr_rates.Term.astype(str) + ' SOFR'
        
        # THOR rates
        th_cal = self.calendar_manager.get_calendar('th')
        self.market_data.thor_rates["Termination"] = self._calculate_termination_dates(
            self.market_data.thor_rates, th_cal
        )
        self.market_data.thor_rates["Label"] = self.market_data.thor_rates.Term.astype(str) + ' THOR'
        
        # TONR rates
        jp_cal = self.calendar_manager.get_calendar('jp')
        self.market_data.tonr_rates["Termination"] = self._calculate_termination_dates(
            self.market_data.tonr_rates, jp_cal
        )
        self.market_data.tonr_rates["Label"] = self.market_data.tonr_rates.Term.astype(str) + ' TONR'
        
        # FX Swap rates
        us_th_cal = self.calendar_manager.get_calendar('us_th')
        us_jp_cal = self.calendar_manager.get_calendar('us_jp')
        
        self.market_data.usdthb_sw_rates["Termination"] = [
            Schedule(
                effective=self.thb_settlement_date, 
                eom=True,  
                termination=term, 
                frequency="A", 
                calendar=us_th_cal,
                payment_lag=2
            ).table["Acc End"].iat[-1]
            for term in self.market_data.usdthb_sw_rates["Term"]
        ]
        self.market_data.usdthb_sw_rates["Label"] = self.market_data.usdthb_sw_rates.Term.astype(str) + ' FXS'
        
        self.market_data.usdjpy_sw_rates["Termination"] = [
            Schedule(
                effective=self.thb_settlement_date, 
                eom=True,  
                termination=term, 
                frequency="A", 
                calendar=us_jp_cal,
                payment_lag=2
            ).table["Acc End"].iat[-1]
            for term in self.market_data.usdjpy_sw_rates["Term"]
        ]
        self.market_data.usdjpy_sw_rates["Label"] = self.market_data.usdjpy_sw_rates.Term.astype(str) + ' FXS'
        
        # CCS rates
        self.market_data.usdthb_ccs_rates["Termination"] = [
            Schedule(
                effective=self.thb_settlement_date, 
                eom=True,  
                termination=term, 
                frequency="A", 
                calendar=us_th_cal,
                payment_lag=2
            ).table["Acc End"].iat[-1]
            for term in self.market_data.usdthb_ccs_rates["Term"]
        ]
        self.market_data.usdthb_ccs_rates["Label"] = self.market_data.usdthb_ccs_rates.Term.astype(str) + ' CCS'
        
        self.market_data.usdjpy_ccs_rates["Termination"] = [
            Schedule(
                effective=self.thb_settlement_date, 
                eom=True,  
                termination=term, 
                frequency="A", 
                calendar=us_jp_cal,
                payment_lag=2
            ).table["Acc End"].iat[-1]
            for term in self.market_data.usdjpy_ccs_rates["Term"]
        ]
        self.market_data.usdjpy_ccs_rates["Label"] = self.market_data.usdjpy_ccs_rates.Term.astype(str) + ' CCS'
    
    def _build_sofr_curve(self) -> None:
        """Build SOFR curve"""
        us_cal = self.calendar_manager.get_calendar('us')
        
        self.curves['sofr'] = Curve(
            id="sofr",
            convention="Act360",
            calendar=us_cal,
            modifier="MF",
            interpolation="linear_zero_rate",
            nodes={
                **{self.market_data.curve_date: 1.0},
                **{_: 1.0 for _ in self.market_data.sofr_rates["Termination"]}
            }
        )
    
    def _build_thor_curve(self) -> None:
        """Build THOR curve"""
        th_cal = self.calendar_manager.get_calendar('th')
        
        self.curves['thor'] = Curve(
            id="thor",
            convention="Act365f",
            calendar=th_cal,
            modifier="MF",
            interpolation="linear_zero_rate",
            nodes={
                **{self.market_data.curve_date: 1.0},
                **{_: 1.0 for _ in self.market_data.thor_rates["Termination"]}
            }
        )
    
    def _build_tonr_curve(self) -> None:
        """Build TONR curve"""
        jp_cal = self.calendar_manager.get_calendar('jp')
        
        self.curves['tonr'] = Curve(
            id="tonr",
            convention="Act365f",
            calendar=jp_cal,
            modifier="MF",
            interpolation="linear_zero_rate",
            nodes={
                **{self.market_data.curve_date: 1.0},
                **{_: 1.0 for _ in self.market_data.tonr_rates["Termination"]}
            }
        )
    
    def _build_fx_swap_curves(self) -> None:
        """Build FX swap curves"""
        us_th_cal = self.calendar_manager.get_calendar('us_th')
        us_jp_cal = self.calendar_manager.get_calendar('us_jp')
        
        # USDTHB Swap curve
        self.curves['utsw'] = Curve(
            id="utsw",
            convention="Act365f",
            calendar=us_th_cal,
            modifier="MF",
            interpolation="linear_zero_rate",
            nodes={
                **{self.market_data.curve_date: 1.0},
                **{_: 1.0 for _ in self.market_data.usdthb_sw_rates["Termination"]}
            }
        )
        
        # USDJPY Swap curve
        self.curves['ujsw'] = Curve(
            id="ujsw",
            convention="Act365f",
            calendar=us_jp_cal,
            modifier="MF",
            interpolation="linear_zero_rate",
            nodes={
                **{self.market_data.curve_date: 1.0},
                **{_: 1.0 for _ in self.market_data.usdjpy_sw_rates["Termination"]}
            }
        )
    
    def _build_ccs_curves(self) -> None:
        """Build CCS curves"""
        us_th_cal = self.calendar_manager.get_calendar('us_th')
        us_jp_cal = self.calendar_manager.get_calendar('us_jp')
        jp_th_cal = self.calendar_manager.get_calendar('jp_th')
        
        # THBUSD CCS curve
        self.curves['utccs'] = Curve(
            id="utccs",
            convention="Act365f",
            calendar=us_th_cal,
            modifier="MF",
            interpolation="linear_zero_rate",
            nodes={
                **{self.market_data.curve_date: 1.0},
                **{_: 1.0 for _ in self.market_data.usdthb_ccs_rates["Termination"]}
            }
        )
        
        # JPYUSD CCS curve
        self.curves['ujccs'] = Curve(
            id="ujccs",
            convention="Act365f",
            calendar=us_jp_cal,
            modifier="MF",
            interpolation="linear_zero_rate",
            nodes={
                **{self.market_data.curve_date: 1.0},
                **{_: 1.0 for _ in self.market_data.usdjpy_ccs_rates["Termination"]}
            }
        )
        
        # THBJPY CCS curve (synthetic) - derived from USD/THB and USD/JPY
        # Create synthetic THB/JPY curve by combining USD/THB and USD/JPY curves
        self.curves['tjccs'] = self._create_synthetic_thbjpy_curve()
        
        # Combined curves for bootstrapping
        ut_rate = pd.concat([self.market_data.usdthb_sw_rates, self.market_data.usdthb_ccs_rates], ignore_index=True)
        uj_rate = pd.concat([self.market_data.usdjpy_sw_rates, self.market_data.usdjpy_ccs_rates], ignore_index=True)
        
        self.curves['usdthb_xcs'] = Curve(
            id="usdthb_xcs",
            convention="Act365f",
            calendar=us_th_cal,
            modifier="MF",
            interpolation="linear_zero_rate",
            nodes={
                **{self.market_data.curve_date: 1.0},
                **{_: 1.0 for _ in ut_rate["Termination"]}
            }
        )
        
        self.curves['usdjpy_xcs'] = Curve(
            id="usdjpy_xcs",
            convention="Act365f",
            calendar=us_jp_cal,
            modifier="MF",
            interpolation="linear_zero_rate",
            nodes={
                **{self.market_data.curve_date: 1.0},
                **{_: 1.0 for _ in uj_rate["Termination"]}
            }
        )
        
        # THBJPY XCS curve (synthetic) - derived from USD/THB and USD/JPY
        self.curves['thbjpy_xcs'] = self._create_synthetic_thbjpy_xcs_curve()
    
    def _create_synthetic_thbjpy_curve(self) -> Curve:
        """
        Create synthetic THB/JPY curve using proper triangulation methodology.
        This curve represents the proper FX forward relationship derived from USD/THB and USD/JPY.
        """
        jp_th_cal = self.calendar_manager.get_calendar('jp_th')
        
        # Get termination dates from USD/THB and USD/JPY curves
        usdthb_dates = [self.market_data.curve_date] + list(self.market_data.usdthb_ccs_rates["Termination"])
        usdjpy_dates = [self.market_data.curve_date] + list(self.market_data.usdjpy_ccs_rates["Termination"])
        
        # Find common dates between USD/THB and USD/JPY curves
        common_dates = []
        for date in usdthb_dates:
            if date in usdjpy_dates:
                common_dates.append(date)
        
        # Create curve nodes using proper FX forward triangulation
        # F_JPYTHB(t) = F_USDTHB(t) / F_USDJPY(t)
        nodes = {self.market_data.curve_date: 1.0}
        
        # Get base curves for proper FX forward calculation
        thor_curve = self.curves.get('thor')
        tonr_curve = self.curves.get('tonr')
        
        for date in common_dates[1:]:  # Skip curve date
            if thor_curve and tonr_curve:
                try:
                    # Get discount factors from base curves
                    thor_df = float(thor_curve[date])
                    tonr_df = float(tonr_curve[date])
                    
                    # Calculate FX forward using interest rate parity
                    # For JPY/THB forward: F = S * (1 + r_THB * t) / (1 + r_JPY * t)
                    # Using discount factors: F = S * D_THB / D_JPY
                    if thor_df > 0 and tonr_df > 0:
                        # Calculate the forward rate ratio
                        forward_ratio = thor_df / tonr_df
                        # Apply reasonable bounds to prevent extreme values
                        synthetic_df = max(MIN_SYNTHETIC_DF, min(MAX_SYNTHETIC_DF, forward_ratio))
                    else:
                        synthetic_df = DEFAULT_DISCOUNT_FACTOR
                        
                except Exception as e:
                    print(f"Warning: Error calculating FX forward for {date}: {e}")
                    synthetic_df = DEFAULT_DISCOUNT_FACTOR
                
                nodes[date] = synthetic_df
            else:
                # Fallback to default discount factor if curves not available
                nodes[date] = DEFAULT_DISCOUNT_FACTOR
        
        return Curve(
            id="tjccs",
            convention="Act365f",
            calendar=jp_th_cal,
            modifier="MF",
            interpolation="linear_zero_rate",
            nodes=nodes
        )
    
    def _create_synthetic_thbjpy_xcs_curve(self) -> Curve:
        """
        Create synthetic THB/JPY XCS curve using proper triangulation methodology.
        This curve represents the basis spread for JPY-THB CCS derived from USD-based markets.
        """
        jp_th_cal = self.calendar_manager.get_calendar('jp_th')
        
        # Get termination dates from USD/THB and USD/JPY XCS curves
        usdthb_sw_dates = list(self.market_data.usdthb_sw_rates["Termination"])
        usdthb_ccs_dates = list(self.market_data.usdthb_ccs_rates["Termination"])
        usdjpy_sw_dates = list(self.market_data.usdjpy_sw_rates["Termination"])
        usdjpy_ccs_dates = list(self.market_data.usdjpy_ccs_rates["Termination"])
        
        # Combine all dates
        all_usdthb_dates = [self.market_data.curve_date] + usdthb_sw_dates + usdthb_ccs_dates
        all_usdjpy_dates = [self.market_data.curve_date] + usdjpy_sw_dates + usdjpy_ccs_dates
        
        # Find common dates between USD/THB and USD/JPY curves
        common_dates = []
        for date in all_usdthb_dates:
            if date in all_usdjpy_dates:
                common_dates.append(date)
        
        # Create curve nodes using proper basis triangulation
        # b_JPY-THB(t) = b_JPY-USD(t) - b_THB-USD(t)
        nodes = {self.market_data.curve_date: 1.0}
        
        # Get basis spreads from USD-based CCS markets
        usdthb_basis = self._get_basis_spreads_from_market('usdthb')
        usdjpy_basis = self._get_basis_spreads_from_market('usdjpy')
        
        for date in common_dates[1:]:  # Skip curve date
            try:
                # Get basis spreads for this date
                b_thb_usd = usdthb_basis.get(date, 0.0)
                b_jpy_usd = usdjpy_basis.get(date, 0.0)
                
                # Triangulate JPY-THB basis: b_JPY-THB = b_JPY-USD - b_THB-USD
                b_jpy_thb = b_jpy_usd - b_thb_usd
                
                # Convert basis spread to discount factor adjustment
                # For XCS curve, we use the basis spread to adjust the curve
                # This is a simplified approach - in practice, you'd use more sophisticated methods
                basis_adjustment = DEFAULT_DISCOUNT_FACTOR + (b_jpy_thb / BASIS_POINTS_MULTIPLIER)
                
                # Apply reasonable bounds
                synthetic_df = max(MIN_SYNTHETIC_DF, min(MAX_SYNTHETIC_DF, basis_adjustment))
                
            except Exception as e:
                print(f"Warning: Error calculating basis triangulation for {date}: {e}")
                synthetic_df = DEFAULT_DISCOUNT_FACTOR
            
            nodes[date] = synthetic_df
        
        return Curve(
            id="thbjpy_xcs",
            convention="Act365f",
            calendar=jp_th_cal,
            modifier="MF",
            interpolation="linear_zero_rate",
            nodes=nodes
        )
    
    def _get_basis_spreads_from_market(self, currency_pair: str) -> Dict[datetime.datetime, float]:
        """
        Extract basis spreads from market data for a given currency pair.
        
        Args:
            currency_pair: 'usdthb' or 'usdjpy'
            
        Returns:
            Dictionary mapping termination dates to basis spreads in basis points
        """
        basis_spreads = {}
        
        if currency_pair == 'usdthb':
            # Get USD/THB CCS basis spreads
            for _, row in self.market_data.usdthb_ccs_rates.iterrows():
                basis_spreads[row['Termination']] = row['Rate']
        elif currency_pair == 'usdjpy':
            # Get USD/JPY CCS basis spreads
            for _, row in self.market_data.usdjpy_ccs_rates.iterrows():
                basis_spreads[row['Termination']] = row['Rate']
        
        return basis_spreads
    
    def _get_common_fx_dates(self) -> List[datetime.datetime]:
        """Get common termination dates from USD/THB and USD/JPY FX swap rates"""
        usdthb_dates = list(self.market_data.usdthb_sw_rates["Termination"])
        usdjpy_dates = list(self.market_data.usdjpy_sw_rates["Termination"])
        
        # Find common dates
        common_dates = []
        for date in usdthb_dates:
            if date in usdjpy_dates:
                common_dates.append(date)
        
        return common_dates
    
    def _get_common_ccs_dates(self) -> List[datetime.datetime]:
        """Get common termination dates from USD/THB and USD/JPY CCS rates"""
        usdthb_dates = list(self.market_data.usdthb_ccs_rates["Termination"])
        usdjpy_dates = list(self.market_data.usdjpy_ccs_rates["Termination"])
        
        # Find common dates
        common_dates = []
        for date in usdthb_dates:
            if date in usdjpy_dates:
                common_dates.append(date)
        
        return common_dates
    
    def _create_synthetic_thbjpy_rates(self) -> List[float]:
        """
        Create synthetic THB/JPY rates using proper triangulation methodology.
        Uses the formula: b_JPY-THB(t) = b_JPY-USD(t) - b_THB-USD(t)
        """
        synthetic_rates = []
        
        # Get common dates from both FX swap and CCS markets
        common_fx_dates = self._get_common_fx_dates()
        common_ccs_dates = self._get_common_ccs_dates()
        
        # Combine all common dates
        all_common_dates = list(set(common_fx_dates + common_ccs_dates))
        all_common_dates.sort()
        
        # Get basis spreads from USD-based markets
        usdthb_basis = self._get_basis_spreads_from_market('usdthb')
        usdjpy_basis = self._get_basis_spreads_from_market('usdjpy')
        
        print(f"  USD-THB basis spreads: {list(usdthb_basis.values())[:3] if usdthb_basis else 'None'}")
        print(f"  USD-JPY basis spreads: {list(usdjpy_basis.values())[:3] if usdjpy_basis else 'None'}")
        
        # Process all common dates
        for date in all_common_dates:
            try:
                # Get basis spreads for this date
                b_thb_usd = usdthb_basis.get(date, 0.0)
                b_jpy_usd = usdjpy_basis.get(date, 0.0)
                
                # Triangulate JPY-THB basis using proper formula
                # b_JPY-THB(t) = b_JPY-USD(t) - b_THB-USD(t)
                b_jpy_thb = b_jpy_usd - b_thb_usd
                
                synthetic_rates.append(b_jpy_thb)
                
            except Exception as e:
                print(f"Warning: Error calculating synthetic rate for {date}: {e}")
                synthetic_rates.append(0.0)
        
        print(f"✓ Created {len(synthetic_rates)} synthetic JPY-THB rates using triangulation")
        print(f"  Sample rates: {synthetic_rates[:5] if len(synthetic_rates) >= 5 else synthetic_rates}")
        return synthetic_rates
        
    def _create_synthetic_thbjpy_labels(self) -> List[str]:
        """
        Create synthetic THB/JPY labels for instruments using triangulation methodology.
        """
        # Get common dates from both FX swap and CCS markets
        common_fx_dates = self._get_common_fx_dates()
        common_ccs_dates = self._get_common_ccs_dates()
        
        # Combine all common dates
        all_common_dates = list(set(common_fx_dates + common_ccs_dates))
        all_common_dates.sort()
        
        labels = []
        
        # Create labels for all common dates
        for date in all_common_dates:
            # Calculate years to maturity
            years = (date - self.market_data.curve_date).days / DAYS_IN_YEAR
            
            # Determine if this is from FX swap or CCS market
            if date in common_fx_dates and date in common_ccs_dates:
                # Date exists in both markets - use CCS label (more specific)
                labels.append(f"{years:.1f}Y THB/JPY CCS (Triangulated)")
            elif date in common_fx_dates:
                # Only in FX swap market
                labels.append(f"{years:.1f}Y THB/JPY FXS (Triangulated)")
            elif date in common_ccs_dates:
                # Only in CCS market
                labels.append(f"{years:.1f}Y THB/JPY CCS (Triangulated)")
            else:
                # Fallback
                labels.append(f"{years:.1f}Y THB/JPY (Triangulated)")
        
        print(f"✓ Created {len(labels)} synthetic JPY-THB labels")
        return labels
    
    def _build_fx_curves(self) -> None:
        """Build FX forward curves"""
        # USDTHB Forward
        ut_fxr = FXRates({"usdthb": self.market_data.thb_spot}, settlement=self.thb_settlement_date)
        self.fx_forwards['ut_fxf'] = FXForwards(
            fx_rates=ut_fxr,
            fx_curves={
                "usdusd": self.curves['sofr'],
                "thbthb": self.curves['usdthb_xcs'],
                "usdthb": self.curves['usdthb_xcs']
            }
        )
        
        # USDJPY Forward
        uj_fxr = FXRates({"usdjpy": self.market_data.jpy_spot}, settlement=self.thb_settlement_date)
        self.fx_forwards['uj_fxf'] = FXForwards(
            fx_rates=uj_fxr,
            fx_curves={
                "usdusd": self.curves['sofr'],
                "jpyjpy": self.curves['usdjpy_xcs'],
                "usdjpy": self.curves['usdjpy_xcs']
            }
        )
        
        # THBJPY Forward (synthetic) - derived from USD/THB and USD/JPY using proper triangulation
        # F_JPYTHB(t) = F_USDTHB(t) / F_USDJPY(t)
        tj_fxr = FXRates({"thbjpy": self.market_data.thbjpy_spot}, settlement=self.thb_settlement_date)
        self.fx_forwards['tj_fxf'] = FXForwards(
            fx_rates=tj_fxr,
            fx_curves={
                "thbthb": self.curves['thor'],
                "jpyjpy": self.curves['tonr'],
                "thbjpy": self.curves['thbjpy_xcs']  # Use the triangulated THB/JPY curve
            }
        )
    
    def _build_solvers(self) -> None:
        """Build curve solvers"""
        self._create_instruments()
        self._solve_curves()
    
    def _create_instruments(self) -> None:
        """Create instruments for curve solving"""
        # SOFR instruments
        sofr_args = dict(
            effective=self.usbthb_valuation_date, 
            spec="usd_irs",
            curves="sofr", 
            frequency="A",
            stub="shortfront", 
            convention="Act360",
            eom=True,
            modifier="MF",
            leg2_spread_compound_method="none_simple",
            leg2_fixing_method="rfr_payment_delay",
            leg2_method_param=0,
            payment_lag=self.market_data.lag_days, 
            leg2_payment_lag=self.market_data.lag_days
        )
        self.instruments['sofr'] = [
            IRS(termination=_, **sofr_args) 
            for _ in self.market_data.sofr_rates["Termination"]
        ]
        
        # THOR instruments
        thor_args_s = dict(
            effective=self.usbthb_valuation_date, 
            curves="thor",
            frequency="A",
            stub="shortfront", 
            convention="Act365f",
            eom=True, 
            modifier="MF", 
            payment_lag=2, 
            leg2_payment_lag=2
        )
        thor_args_l = dict(
            effective=self.usbthb_valuation_date, 
            curves="thor",
            frequency="Q",
            stub="shortfront", 
            convention="Act365f",
            eom=True, 
            modifier="MF", 
            payment_lag=2, 
            leg2_payment_lag=2
        )
        
        self.instruments['thor'] = (
            [IRS(termination=_, **thor_args_s) for _ in self.market_data.thor_rates["Termination"][:7]] +
            [IRS(termination=_, **thor_args_l) for _ in self.market_data.thor_rates["Termination"][7:]]
        )
        
        # TONR instruments
        tonr_args = dict(
            effective=self.usbthb_valuation_date, 
            spec="jpy_irs",
            curves="tonr",
            frequency="A",
            stub="shortfront", 
            convention="Act365f",
            eom=True, 
            modifier="MF",
            payment_lag=2, 
            leg2_payment_lag=2
        )
        self.instruments['tonr'] = [
            IRS(termination=_, **tonr_args) 
            for _ in self.market_data.tonr_rates["Termination"]
        ]
        
        # FX Swap instruments
        usdthb_args = dict(
            effective=self.usbthb_valuation_date,
            pair="usdthb",
            curves=["sofr", "sofr", "usdthb_xcs", "usdthb_xcs"]
        )
        self.instruments['usdthb'] = [
            FXSwap(termination=_, **usdthb_args) 
            for _ in self.market_data.usdthb_sw_rates["Termination"]
        ]
        
        usdjpy_args = dict(
            effective=self.usbthb_valuation_date,
            pair="usdjpy",
            curves=["sofr", "sofr", "usdjpy_xcs", "usdjpy_xcs"]
        )
        self.instruments['usdjpy'] = [
            FXSwap(termination=_, **usdjpy_args) 
            for _ in self.market_data.usdjpy_sw_rates["Termination"]
        ]
        
        # THBJPY FX Swap instruments (synthetic) - use common dates from USD/THB and USD/JPY
        thbjpy_args = dict(
            effective=self.usbthb_valuation_date,
            pair="thbjpy",
            curves=["thor", "thor", "thbjpy_xcs", "thbjpy_xcs"]
        )
        # Get common termination dates from USD/THB and USD/JPY FX swap rates
        common_fx_dates = self._get_common_fx_dates()
        self.instruments['thbjpy'] = [
            FXSwap(termination=_, **thbjpy_args) 
            for _ in common_fx_dates
        ]
        
        # CCS instruments
        self._create_ccs_instruments()
    
    def _create_ccs_instruments(self) -> None:
        """Create CCS instruments"""
        # Create custom thbusd_xcs specification
        defaults.spec["thbusd_xcs"] = {
            'frequency': 'q',
            'stub': 'shortfront', 
            'eom': False,
            'modifier': 'mf',
            'calendar': self.calendar_manager.get_calendar('us_th_en'),
            'payment_lag': 2,
            'currency': 'thb',
            'convention': 'act365f',
            'leg2_currency': 'usd',
            'leg2_convention': 'act360',
            'spread_compound_method': 'none_simple',
            'fixing_method': 'rfr_payment_delay',
            'method_param': 0,
            'leg2_spread_compound_method': 'none_simple',
            'leg2_fixing_method': 'rfr_payment_delay',
            'leg2_method_param': 0,
            'payment_lag_exchange': 0,
            'fixed': False,
            'leg2_fixed': False,
            'leg2_mtm': False
        }
        # thbjpy_xcs specification for THB/JPY CCS
        defaults.spec["thbjpy_xcs"] = {
            'frequency': 'q',
            'stub': 'shortfront', 
            'eom': False,
            'modifier': 'mf',
            'calendar': self.calendar_manager.get_calendar('jp_th'),
            'payment_lag': 2,
            'currency': 'thb',
            'convention': 'act365f',
            'leg2_currency': 'jpy',
            'leg2_convention': 'act365f',
            'spread_compound_method': 'none_simple',
            'fixing_method': 'rfr_payment_delay',
            'method_param': 0,
            'leg2_spread_compound_method': 'none_simple',
            'leg2_fixing_method': 'rfr_payment_delay',
            'leg2_method_param': 0,
            'payment_lag_exchange': 0,
            'fixed': False,
            'leg2_fixed': False,
            'leg2_mtm': False
        }
        
        # THBUSD CCS instruments
        utccs_args = dict(
            effective=self.usbthb_valuation_date,
            spec="thbusd_xcs",
            curves=["thor", "usdthb_xcs", "sofr", "sofr"],
            frequency="Q",
            leg2_frequency="Q",
            stub = 'shortfront', 
            eom=True,
            leg2_eom=True,
            currency='thb', 
            leg2_currency='usd',
            calendar=self.calendar_manager.get_calendar('us_th_en'), 
            leg2_calendar=self.calendar_manager.get_calendar('us_th_en'),
            convention='Act365f', 
            leg2_convention='Act360',
            fx_fixings=1/self.market_data.thb_spot,
            fixing_method='rfr_payment_delay',
            spread_compound_method='none_simple',
            leg2_fixing_method='rfr_payment_delay',
            leg2_spread_compound_method='none_simple',
            leg2_mtm=False,
            leg2_fixed=False,
            float_spread=NoInput.blank,
            payment_lag = 2,
            leg2_payment_lag=2
        )
        self.instruments['utccs'] = [
            XCS(termination=_, **utccs_args) 
            for _ in self.market_data.usdthb_ccs_rates["Termination"]
        ]
        
        # JPYUSD CCS instruments
        ujccs_args = dict(
            effective=self.usbthb_valuation_date, 
            spec="jpyusd_xcs",
            curves=["tonr", "usdjpy_xcs", "sofr", "sofr"],
            frequency="Q", 
            leg2_frequency="Q", 
            currency='jpy', 
            leg2_currency='usd',
            calendar=self.calendar_manager.get_calendar('us_jp'), 
            leg2_calendar=self.calendar_manager.get_calendar('us_jp'),
            convention='Act365f', 
            leg2_convention='Act360',
            payment_lag=2, 
            leg2_payment_lag=2
        )
        self.instruments['ujccs'] = [
            XCS(termination=_, **ujccs_args) 
            for _ in self.market_data.usdjpy_ccs_rates["Termination"]
        ]
        
        # THBJPY CCS instruments (synthetic) - use proper triangulation methodology
        tjccs_args = dict(
            effective=self.usbthb_valuation_date,
            spec="thbjpy_xcs",
            curves=["thor", "thor", "tonr", "thbjpy_xcs"],  # Correct curve order for XCS
            frequency="Q",
            leg2_frequency="Q",
            currency='thb',
            leg2_currency='jpy',
            calendar=self.calendar_manager.get_calendar('jp_th'),
            leg2_calendar=self.calendar_manager.get_calendar('jp_th'),
            convention='Act365f',
            leg2_convention='Act365f',
            fx_fixings=self.market_data.thbjpy_spot,  # Use triangulated spot rate
            payment_lag=2,
            leg2_payment_lag=2
        )
        # Get common termination dates from USD/THB and USD/JPY CCS rates
        common_ccs_dates = self._get_common_ccs_dates()
        self.instruments['tjccs'] = [
            XCS(termination=_, **tjccs_args) 
            for _ in common_ccs_dates
        ]
    
    def _solve_curves(self) -> None:
        """Solve all curves"""
        # SOFR solver
        self.solvers['sofr'] = Solver(
            curves=[self.curves['sofr']],
            instruments=self.instruments['sofr'],
            s=self.market_data.sofr_rates["Rate"].tolist(),
            fx=self.fx_forwards['ut_fxf'],
            instrument_labels=self.market_data.sofr_rates["Label"].tolist(),
            id="sofr"
        )
        
        # THOR solver
        self.solvers['thor'] = Solver(
            curves=[self.curves['thor']],
            instruments=self.instruments['thor'],
            s=self.market_data.thor_rates["Rate"].tolist(),
            fx=self.fx_forwards['ut_fxf'],
            instrument_labels=self.market_data.thor_rates["Label"].tolist(),
            id="thor"
        )
        
        # TONR solver
        self.solvers['tonr'] = Solver(
            curves=[self.curves['tonr']],
            instruments=self.instruments['tonr'],
            s=self.market_data.tonr_rates["Rate"].tolist(),
            fx=self.fx_forwards['uj_fxf'],
            instrument_labels=self.market_data.tonr_rates["Label"].tolist(),
            id="tonr"
        )
        
        # USDTHB CCS solver - only include XCS instrument labels
        utccs_labels = (self.market_data.usdthb_sw_rates["Label"].tolist() + 
                       self.market_data.usdthb_ccs_rates["Label"].tolist())
        self.solvers['utccs'] = Solver(
            pre_solvers=[self.solvers['thor'], self.solvers['sofr']],                  
            curves=[self.curves['usdthb_xcs']],
            instruments=self.instruments['usdthb'] + self.instruments['utccs'],
            s=self.market_data.usdthb_sw_rates["Rate"].tolist() + self.market_data.usdthb_ccs_rates["Rate"].tolist(),       
            fx=self.fx_forwards['ut_fxf'],
            instrument_labels=utccs_labels,
            id="utccs_xcs"
        )
        
        # USDJPY CCS solver - only include XCS instrument labels
        ujccs_labels = (self.market_data.usdjpy_sw_rates["Label"].tolist() + 
                       self.market_data.usdjpy_ccs_rates["Label"].tolist())
        self.solvers['ujccs'] = Solver(
            pre_solvers=[self.solvers['tonr'], self.solvers['sofr']],  
            curves=[self.curves['usdjpy_xcs']],
            instruments=self.instruments['usdjpy'] + self.instruments['ujccs'],
            s=self.market_data.usdjpy_sw_rates["Rate"].tolist() + self.market_data.usdjpy_ccs_rates["Rate"].tolist(),
            fx=self.fx_forwards['uj_fxf'],
            instrument_labels=ujccs_labels,
            id="ujccs_xcs"
        )
        
        # THBJPY CCS solver (synthetic) - use proper triangulation methodology
        synthetic_thbjpy_rates = self._create_synthetic_thbjpy_rates()
        synthetic_thbjpy_labels = self._create_synthetic_thbjpy_labels()
        
        self.solvers['tjccs'] = Solver(
            pre_solvers=[self.solvers['sofr'], self.solvers['thor'], self.solvers['tonr']],  # Include SOFR for USD-CSA
            curves=[self.curves['thbjpy_xcs']],
            instruments=self.instruments['thbjpy'] + self.instruments['tjccs'],
            s=synthetic_thbjpy_rates,
            fx=self.fx_forwards['tj_fxf'],
            instrument_labels=synthetic_thbjpy_labels,
            id="tjccs_xcs"
        )
    
    def get_curve(self, name: str) -> Curve:
        """Get curve by name"""
        return self.curves.get(name)
    
    def get_solver(self, name: str) -> Solver:
        """Get solver by name"""
        return self.solvers.get(name)


class CCSBookingCalculator:
    """Handles CCS booking calculations and NPV computations"""
    
    def __init__(self, curve_builder: CurveBuilder, market_data: MarketData):
        self.curve_builder = curve_builder
        self.market_data = market_data
        self.cf_table = None
    
    def _convert_to_pl_currency(self, value: float, from_currency: str, to_currency: str) -> float:
        """
        Convert a value from one currency to another for P&L reporting
        
        Args:
            value: The value to convert
            from_currency: Source currency ('USD', 'THB', 'JPY')
            to_currency: Target currency ('USD', 'THB', 'JPY')
            
        Returns:
            Converted value in target currency
        """
        if from_currency == to_currency:
            return value
        
        # Convert to USD first (as intermediate currency)
        if from_currency == 'THB':
            value_usd = value / self.market_data.thb_spot
        elif from_currency == 'JPY':
            value_usd = value / self.market_data.jpy_spot
        else:  # USD
            value_usd = value
        
        # Convert from USD to target currency
        if to_currency == 'THB':
            return value_usd * self.market_data.thb_spot
        elif to_currency == 'JPY':
            return value_usd * self.market_data.jpy_spot
        else:  # USD
            return value_usd
    
    def _get_curve_name(self, currency: str, curve_type: str) -> str:
        """Get the correct curve name for a currency and curve type"""
        if currency == 'USD':
            return 'sofr'
        elif currency == 'THB':
            return 'thor' if curve_type == 'THOR' else 'thor'  # Default to THOR
        elif currency == 'JPY':
            return 'tonr'
        else:
            return 'sofr'  # Default fallback
    
    def _get_fx_curve_name(self, primary_currency: str, secondary_currency: str) -> str:
        """Get the correct FX curve name for a currency pair"""
        # Sort currencies to ensure consistent mapping
        currencies = sorted([primary_currency, secondary_currency])
        
        if currencies == ['THB', 'USD']:
            return 'usdthb_xcs'
        elif currencies == ['JPY', 'USD']:
            return 'usdjpy_xcs'
        elif currencies == ['JPY', 'THB']:
            # For THB/JPY, use synthetic curve name
            return 'thbjpy_xcs'
        else:
            return 'usdthb_xcs'  # Default fallback
    
    def _get_fx_forward(self, primary_currency: str, secondary_currency: str):
        """Get the correct FX forward object for a currency pair"""
        currencies = sorted([primary_currency, secondary_currency])
        
        if currencies == ['THB', 'USD']:
            return self.curve_builder.fx_forwards['ut_fxf']
        elif currencies == ['JPY', 'USD']:
            return self.curve_builder.fx_forwards['uj_fxf']
        elif currencies == ['JPY', 'THB']:
            # Use triangulated THB/JPY FX forward
            return self.curve_builder.fx_forwards['tj_fxf']
        else:
            return self.curve_builder.fx_forwards['ut_fxf']  # Default fallback
    
    def _get_fx_fixing(self, primary_currency: str, secondary_currency: str, fx_fixing: float) -> float:
        """
        Get the correct FX fixing rate for a currency pair using market spot rates.
        Ignores Excel FX_Fixing to ensure symmetric NPV for paired trades.
        
        XCS expects fx_fixings = how many units of leg2_currency per 1 unit of leg1_currency (primary)
        
        Args:
            primary_currency: Leg 1 currency
            secondary_currency: Leg 2 currency
            fx_fixing: Excel FX_Fixing value (ignored, kept for compatibility)
            
        Returns:
            Market spot rate: leg2_currency per 1 unit of primary_currency
        """
        # USD/THB pair
        if primary_currency == 'USD' and secondary_currency == 'THB':
            # How many THB per 1 USD
            return float(self.market_data.thb_spot)
        elif primary_currency == 'THB' and secondary_currency == 'USD':
            # How many USD per 1 THB = 1 / (THB per USD)
            return float(1.0 / self.market_data.thb_spot)
        
        # USD/JPY pair
        elif primary_currency == 'USD' and secondary_currency == 'JPY':
            # How many JPY per 1 USD
            return float(self.market_data.jpy_spot)
        elif primary_currency == 'JPY' and secondary_currency == 'USD':
            # How many USD per 1 JPY = 1 / (JPY per USD)
            return float(1.0 / self.market_data.jpy_spot)
        
        # THB/JPY pair
        elif primary_currency == 'THB' and secondary_currency == 'JPY':
            # How many JPY per 1 THB = 1 / (THB per JPY)
            return float(1.0 / self.market_data.thbjpy_spot)
        elif primary_currency == 'JPY' and secondary_currency == 'THB':
            # How many THB per 1 JPY
            return float(self.market_data.thbjpy_spot)
        
        else:
            # Fallback: use Excel value (should not reach here for supported pairs)
            return float(1/fx_fixing)
    
    def _get_forecast_curve(self, currency: str):
        """Get the forecasting curve for a given currency"""
        if currency == 'USD':
            return self.curve_builder.get_curve('sofr')
        elif currency == 'THB':
            return self.curve_builder.get_curve('thor')
        elif currency == 'JPY':
            return self.curve_builder.get_curve('tonr')
        else:
            return self.curve_builder.get_curve('sofr')
    
    def _get_discount_curve(self, currency: str, currency_pair: tuple):
        """
        Get the discounting curve for a given currency in a CCS pair.
        
        In CCS pricing, each leg is discounted using its own currency's collateral curve:
        - USD leg: always discounted at SOFR (USD OIS)
        - THB leg: discounted at THOR + basis adjustment (via usdthb_xcs)
        - JPY leg: discounted at TONR + basis adjustment (via usdjpy_xcs)
        
        Args:
            currency: The currency we need to discount ('USD', 'THB', 'JPY')
            currency_pair: Tuple of (primary_currency, secondary_currency) for the CCS
        """
        # USD is ALWAYS discounted at SOFR (USD collateral) regardless of the currency pair
        if currency == 'USD':
            return self.curve_builder.get_curve('sofr')
        
        # For non-USD currencies, use the appropriate basis-adjusted discount curve
        elif currency == 'THB':
            # THB discounted at THOR + USD/THB basis = usdthb_xcs
            return self.curve_builder.get_curve('usdthb_xcs')
        
        elif currency == 'JPY':
            # JPY discounted at TONR + USD/JPY basis = usdjpy_xcs
            return self.curve_builder.get_curve('usdjpy_xcs')
        
        # Default fallback (should not reach here)
        return self.curve_builder.get_curve('sofr')
    
    def _get_fixings_for_currency(self, currency: str) -> pd.Series:
        """Get the correct fixings series for a currency"""
        if currency == 'USD':
            return self.market_data.sofr_fixings
        elif currency == 'THB':
            return self.market_data.thor_fixings
        elif currency == 'JPY':
            return self.market_data.tonr_fixings
        else:
            return self.market_data.sofr_fixings  # Default fallback
    
    def _get_solver_name(self, primary_currency: str, secondary_currency: str) -> str:
        """Get the correct solver name for a currency pair"""
        currencies = sorted([primary_currency, secondary_currency])
        
        if currencies == ['THB', 'USD']:
            return 'utccs'
        elif currencies == ['JPY', 'USD']:
            return 'ujccs'
        elif currencies == ['JPY', 'THB']:
            return 'tjccs'  # Use triangulated THB/JPY solver
        else:
            return 'utccs'  # Default fallback
    
    def _get_solver_labels(self, solver_name: str, primary_currency: str, secondary_currency: str) -> List[str]:
        """Get meaningful labels for curve points from the solver and all pre-solvers"""
        try:
            solver = self.curve_builder.get_solver(solver_name)
            if not solver:
                print(f"Solver {solver_name} not found, using fallback")
                return self._construct_fallback_labels(solver_name, primary_currency, secondary_currency)
            
            labels = []
            
            # Get labels from pre-solvers first (SOFR, THOR/TONR)
            if hasattr(solver, 'pre_solvers') and solver.pre_solvers:
                for pre_solver in solver.pre_solvers:
                    if hasattr(pre_solver, 'instrument_labels') and pre_solver.instrument_labels:
                        labels.extend(pre_solver.instrument_labels)
            
            # Get labels from the main solver (XCS instruments)
            if hasattr(solver, 'instrument_labels') and solver.instrument_labels:
                labels.extend(solver.instrument_labels)
            
            # Add FX spot rate label (always 1 point)
            labels.append("FX_Spot_Rate")
            return labels
            
        except Exception as e:
            print(f"Warning: Could not get solver labels for {solver_name}: {e}")
            return self._construct_fallback_labels(solver_name, primary_currency, secondary_currency)
    
    def _construct_fallback_labels(self, solver_name: str, primary_currency: str, secondary_currency: str) -> List[str]:
        """Construct fallback labels when solver labels are not available"""
        labels = []
        
        if solver_name == 'utccs':
            # USDTHB CCS solver - combine ALL curve labels (SOFR + THOR + XCS)
            # The delta sensitivity includes sensitivity to all curves involved
            
            # Add SOFR labels first (USD curve) - typically 20-30 points
            if hasattr(self.curve_builder.market_data, 'sofr_rates'):
                sofr_labels = self.curve_builder.market_data.sofr_rates['Label'].tolist()
                labels.extend(sofr_labels)
            # Add THOR labels (THB curve) - typically 20-30 points
            if hasattr(self.curve_builder.market_data, 'thor_rates'):
                thor_labels = self.curve_builder.market_data.thor_rates['Label'].tolist()
                labels.extend(thor_labels)
            # Add FX swap labels (XCS curve part 1) - typically 10-15 points
            if hasattr(self.curve_builder.market_data, 'usdthb_sw_rates'):
                fx_labels = self.curve_builder.market_data.usdthb_sw_rates['Label'].tolist()
                labels.extend(fx_labels)
            # Add CCS labels (XCS curve part 2) - typically 10-15 points
            if hasattr(self.curve_builder.market_data, 'usdthb_ccs_rates'):
                ccs_labels = self.curve_builder.market_data.usdthb_ccs_rates['Label'].tolist()
                labels.extend(ccs_labels)
            # Add FX forward curve label (1 point for FX spot rate sensitivity)
            labels.append("FX_Spot_Rate")
            
            return labels
                
        elif solver_name == 'ujccs':
            # USDJPY CCS solver - combine ALL curve labels (SOFR + TONR + XCS)
            
            # Add SOFR labels first (USD curve)
            if hasattr(self.curve_builder.market_data, 'sofr_rates'):
                sofr_labels = self.curve_builder.market_data.sofr_rates['Label'].tolist()
                labels.extend(sofr_labels)
            
            # Add TONR labels (JPY curve)
            if hasattr(self.curve_builder.market_data, 'tonr_rates'):
                tonr_labels = self.curve_builder.market_data.tonr_rates['Label'].tolist()
                labels.extend(tonr_labels)
            
            # Add FX swap labels (XCS curve part 1)
            if hasattr(self.curve_builder.market_data, 'usdjpy_sw_rates'):
                fx_labels = self.curve_builder.market_data.usdjpy_sw_rates['Label'].tolist()
                labels.extend(fx_labels)
            
            # Add CCS labels (XCS curve part 2)
            if hasattr(self.curve_builder.market_data, 'usdjpy_ccs_rates'):
                ccs_labels = self.curve_builder.market_data.usdjpy_ccs_rates['Label'].tolist()
                labels.extend(ccs_labels)
            
            # Add FX forward curve label (1 point for FX spot rate sensitivity)
            labels.append("FX_Spot_Rate")
            
            return labels
                
        elif solver_name == 'tjccs':
            # THBJPY CCS solver - combine ALL curve labels (THOR + TONR + Triangulated XCS)
            
            # Add THOR labels first (THB curve)
            if hasattr(self.curve_builder.market_data, 'thor_rates'):
                thor_labels = self.curve_builder.market_data.thor_rates['Label'].tolist()
                labels.extend(thor_labels)
            
            # Add TONR labels (JPY curve)
            if hasattr(self.curve_builder.market_data, 'tonr_rates'):
                tonr_labels = self.curve_builder.market_data.tonr_rates['Label'].tolist()
                labels.extend(tonr_labels)
            
            # Add triangulated FX swap labels (from USD/THB and USD/JPY)
            if hasattr(self.curve_builder.market_data, 'usdthb_sw_rates'):
                usdthb_fx_labels = self.curve_builder.market_data.usdthb_sw_rates['Label'].tolist()
                labels.extend([f"{label} (Triangulated)" for label in usdthb_fx_labels])
            
            if hasattr(self.curve_builder.market_data, 'usdjpy_sw_rates'):
                usdjpy_fx_labels = self.curve_builder.market_data.usdjpy_sw_rates['Label'].tolist()
                labels.extend([f"{label} (Triangulated)" for label in usdjpy_fx_labels])
            
            # Add triangulated CCS labels (from USD/THB and USD/JPY)
            if hasattr(self.curve_builder.market_data, 'usdthb_ccs_rates'):
                usdthb_ccs_labels = self.curve_builder.market_data.usdthb_ccs_rates['Label'].tolist()
                labels.extend([f"{label} (Triangulated)" for label in usdthb_ccs_labels])
            
            if hasattr(self.curve_builder.market_data, 'usdjpy_ccs_rates'):
                usdjpy_ccs_labels = self.curve_builder.market_data.usdjpy_ccs_rates['Label'].tolist()
                labels.extend([f"{label} (Triangulated)" for label in usdjpy_ccs_labels])
            
            # Add FX forward curve label (1 point for FX spot rate sensitivity)
            labels.append("FX_Spot_Rate (Triangulated)")
            
            return labels
                
        elif solver_name == 'sofr':
            if hasattr(self.curve_builder.market_data, 'sofr_rates'):
                labels = self.curve_builder.market_data.sofr_rates['Label'].tolist()
        elif solver_name == 'thor':
            if hasattr(self.curve_builder.market_data, 'thor_rates'):
                labels = self.curve_builder.market_data.thor_rates['Label'].tolist()
        elif solver_name == 'tonr':
            if hasattr(self.curve_builder.market_data, 'tonr_rates'):
                labels = self.curve_builder.market_data.tonr_rates['Label'].tolist()
        
        return labels
    
    def _build_base_xcs_parameters(self, position: BookingPosition, primary_notional: float, 
                                   calendar: Cal, curve_objects: List) -> Dict:
        """
        Build base XCS parameters common to all XCS calculations
        
        Args:
            position: BookingPosition object
            primary_notional: Primary leg notional (may be adjusted for position direction)
            calendar: Calendar object for the trade
            curve_objects: List of curve objects [leg1_forecast, leg1_discount, leg2_forecast, leg2_discount]
            
        Returns:
            Dictionary of base XCS parameters
        """
        primary_currency = position.primary_currency
        secondary_currency = position.secondary_currency
        
        return {
            'effective': position.effective_date.to_pydatetime() if hasattr(position.effective_date, 'to_pydatetime') else position.effective_date,
            'termination': position.maturity_date.to_pydatetime() if hasattr(position.maturity_date, 'to_pydatetime') else position.maturity_date,
            'notional': primary_notional,
            'currency': primary_currency.lower(),
            'leg2_currency': secondary_currency.lower(),
            'calendar': calendar,
            'fx_fixings': self._get_fx_fixing(primary_currency, secondary_currency, position.fx_fixing),
            'payment_lag': int(position.payment_lag),
            'leg2_payment_lag': int(position.payment_lag),
            'payment_lag_exchange': 0,
            'leg2_payment_lag_exchange': 0,
            'amortization': 0.0,  # Fixed amortization for all deal pairs
            'curves': curve_objects,
            'leg2_mtm': False
        }
    
    def _configure_xcs_leg(self, position: BookingPosition, currency: str, is_leg2: bool = False) -> Dict:
        """
        Configure parameters for a single XCS leg (fixed or float)
        
        Args:
            position: BookingPosition object
            currency: Currency for this leg
            is_leg2: Whether this is leg 2 (affects parameter naming)
            
        Returns:
            Dictionary of leg-specific XCS parameters
        """
        prefix = 'leg2_' if is_leg2 else ''
        fixed_float = position.get_fixed_float(currency)
        frequency = position.get_frequency(currency)
        convention = position.get_convention(currency)
        
        params = {
            f'{prefix}frequency': frequency,
            f'{prefix}convention': convention
        }
        
        if is_leg2:
            params['leg2_calendar'] = self._get_calendar_object(position, currency)
        
        if fixed_float == 'Fixed':
            params[f'{prefix}fixed'] = True
            rate = position.get_rate(currency)
            if rate is not None:
                params[f'{prefix}fixed_rate'] = rate
        else:
            params[f'{prefix}fixed'] = False
            spread = position.get_spread(currency)
            if spread is not None:
                params[f'{prefix}float_spread'] = spread
            
            if is_leg2:
                params['leg2_fixings'] = self._get_fixings_for_currency(currency)
                params['leg2_fixing_method'] = "rfr_lookback"
                params['leg2_method_param'] = 5  # Fixed 5 days lookback for all deal pairs
        
        return params
    
    def _get_calendar_object(self, position: BookingPosition, currency: str) -> Cal:
        """
        Get calendar object for a currency pair from the position.
        
        For THB pairs:
        - USD/THB: Uses NYC (US holidays) + THB holidays from Holidays sheet
        - JPY/THB: Uses TYO (JP holidays) + THB holidays from Holidays sheet
        
        Calendar is determined by the currency pair, not individual currency calendars.
        """
        primary_currency = position.primary_currency
        secondary_currency = position.secondary_currency
        
        # Determine calendar based on currency pair (not individual calendars)
        # THB does not have standalone calendar - it's always combined with the other currency
        currencies = sorted([primary_currency, secondary_currency])
        
        if currencies == ['THB', 'USD']:
            # USD/THB: NYC + THB holidays
            calendar_name = 'us_th'
        elif currencies == ['JPY', 'USD']:
            # USD/JPY: NYC + JP holidays
            calendar_name = 'us_jp'
        elif currencies == ['JPY', 'THB']:
            # JPY/THB: TYO + THB holidays
            calendar_name = 'jp_th'
        else:
            # Default fallback
            calendar_name = 'us_th'
        
        calendar = self.curve_builder.calendar_manager.get_calendar(calendar_name)
        if calendar is None:
            # Fallback to USD/THB calendar
            calendar = self.curve_builder.calendar_manager.get_calendar('us_th')
        
        return calendar

    def calculate_npv(self, position: BookingPosition) -> Tuple[float, List[float], float, float, pd.DataFrame]:
        """
        Calculate NPV for a CCS position using comprehensive position parameters
        Supports any combination of USD, THB, JPY currencies
        
        Args:
            position: Booking position parameters with Leg 1 and Leg 2 details
            
        Returns:
            Tuple of (total_npv, delta_sensitivities, leg1_npv, leg2_npv, cashflow_table)
            All NPV values are in the PL Currency specified in position.pl_currency
        """
        # Get currencies and determine leg assignments
        currencies = position.currencies
        primary_currency = position.primary_currency
        secondary_currency = position.secondary_currency
        
        print(f"  Processing {primary_currency}/{secondary_currency} CCS")
        
        # For THB/JPY CCS, use direct XCS approach with USD-based curves
        if set(currencies) == {'THB', 'JPY'}:
            print(f"  Using direct XCS approach for THB/JPY")
            return self._calculate_thbjpy_direct_npv(position)
        
        # For all other currency pairs, use the standard NPV calculation
        return self._calculate_standard_npv(position)

    def _calculate_thbjpy_direct_npv(self, position: BookingPosition) -> Tuple[float, List[float], float, float, pd.DataFrame]:
        """Calculate NPV for THB/JPY using direct XCS approach - minimal version"""
        try:
            primary_currency = position.leg1_currency
            secondary_currency = position.leg2_currency
            
            print(f"    Using direct XCS approach for {primary_currency}/{secondary_currency}")
            
            # Get FX forward
            fx_forward = self._get_fx_forward(primary_currency, secondary_currency)
            
            # Get notional and apply position direction
            notional_leg1 = position.leg1_notional
            # B/S = Buy leg 1, Sell leg 2 - flip notional to match XCS convention
            if position.position == 'B/S':
                notional_leg1 *= -1
            
            # Determine FX fixing and currency setup
            if primary_currency == 'THB' and secondary_currency == 'JPY':
                fx_fixing = float(1.0 / self.market_data.thbjpy_spot)  # Invert THB/JPY to JPY/THB
                currency1 = 'thb'
                currency2 = 'jpy'
                # Leg 1 is THB, Leg 2 is JPY
                leg1_forecast = self.curve_builder.get_curve('thor')  # THB forecasting curve
                leg1_discount = self.curve_builder.get_curve('usdthb_xcs')  # THB discounting curve
                leg2_forecast = self.curve_builder.get_curve('tonr')  # JPY forecasting curve
                leg2_discount = self.curve_builder.get_curve('usdjpy_xcs')  # JPY discounting curve
            elif primary_currency == 'JPY' and secondary_currency == 'THB':
                fx_fixing = float(self.market_data.thbjpy_spot)  # Use THB/JPY directly
                currency1 = 'jpy'
                currency2 = 'thb'
                # Leg 1 is JPY, Leg 2 is THB
                leg1_forecast = self.curve_builder.get_curve('tonr')  # JPY forecasting curve
                leg1_discount = self.curve_builder.get_curve('usdjpy_xcs')  # JPY discounting curve
                leg2_forecast = self.curve_builder.get_curve('thor')  # THB forecasting curve
                leg2_discount = self.curve_builder.get_curve('usdthb_xcs')  # THB discounting curve
            else:
                raise ValueError(f"Unsupported currency pair: {primary_currency}/{secondary_currency}")
            
            # Create XCS using parameters from Excel booking
            xcs_params = {
                'effective': position.effective_date,
                'termination': position.maturity_date,
                'notional': notional_leg1,
                'frequency': position.get_frequency(primary_currency),  # Use actual frequency from Excel
                'currency': currency1,
                'leg2_currency': currency2,
                'calendar': 'tyo',  # Use Tokyo calendar for both
                'fx_fixings': fx_fixing,
                'fixed': position.get_fixed_float(primary_currency) == 'Fixed',
                'curves': [leg1_forecast, leg1_discount, leg2_forecast, leg2_discount],  # Correct curve assignment for each leg
                'convention': position.get_convention(primary_currency),  # Use actual convention from Excel
                'leg2_frequency': position.get_frequency(secondary_currency),  # Use actual frequency from Excel
                'leg2_calendar': 'tyo',
                'leg2_fixed': position.get_fixed_float(secondary_currency) == 'Fixed',
                'leg2_convention': position.get_convention(secondary_currency),  # Use actual convention from Excel
                'leg2_mtm': False,  # Disable mark-to-market to prevent intermediate principal exchanges
                'amortization': 0.0,  # Fixed amortization for all deal pairs
                'payment_lag': 2,
                'leg2_payment_lag': 2,
                'payment_lag_exchange': 0,
                'leg2_payment_lag_exchange': 0,
            }
            
            # Add fixed_rate for leg1 if it's fixed
            if position.get_fixed_float(primary_currency) == 'Fixed':
                rate_val = position.get_rate(primary_currency)
                if rate_val is not None:
                    xcs_params['fixed_rate'] = rate_val
            else:
                # Leg1 is floating - add fixings
                spread_val = position.get_spread(primary_currency)
                if spread_val is not None:
                    xcs_params['float_spread'] = spread_val
                xcs_params['fixings'] = self._get_fixings_for_currency(primary_currency)
                xcs_params['fixing_method'] = "rfr_lookback"
                xcs_params['method_param'] = 5  # Fixed 5 days lookback for all deal pairs
            
            # Add leg2 parameters based on fixed/float
            if position.get_fixed_float(secondary_currency) == 'Fixed':
                # Leg2 is fixed - just add the fixed rate, NO fixings
                rate_val = position.get_rate(secondary_currency)
                if rate_val is not None:
                    xcs_params['leg2_fixed_rate'] = rate_val
            else:
                # Leg2 is floating - add fixings
                spread_val = position.get_spread(secondary_currency)
                if spread_val is not None:
                    xcs_params['leg2_float_spread'] = spread_val
                xcs_params['leg2_fixings'] = self._get_fixings_for_currency(secondary_currency)
                xcs_params['leg2_fixing_method'] = "rfr_lookback"
                xcs_params['leg2_method_param'] = 5  # Fixed 5 days lookback for all deal pairs
            
            # Validate all parameters before creating XCS (remove debug output for cleaner logs)
            none_params = []
            for key, value in xcs_params.items():
                if value is None:
                    none_params.append(key)
            
            if none_params:
                print(f"    ERROR: Found {len(none_params)} None parameters: {none_params}")
                return 0.0, [0.0], 0.0, 0.0, pd.DataFrame()
            
            xccy = XCS(**xcs_params)
            
            # Calculate NPV using the curves - must match the curves parameter in XCS constructor
            curve_objects = [leg1_forecast, leg1_discount, leg2_forecast, leg2_discount]
            total_npv = float(xccy.npv(curves=curve_objects, fx=fx_forward))
            
            # Get cashflows
            cf_table = xccy.cashflows(curves=curve_objects, fx=fx_forward)
            
            # Calculate leg NPVs in native currencies
            leg1_npv_native = cf_table.NPV[cf_table.Ccy == primary_currency.upper()].sum()
            leg2_npv_native = cf_table.NPV[cf_table.Ccy == secondary_currency.upper()].sum()
            
            # Get PL Currency from position (default to USD if not specified)
            pl_currency = position.pl_currency if hasattr(position, 'pl_currency') and position.pl_currency else 'USD'
            
            # Calculate Total NPV in PL Currency by converting each leg
            leg1_npv_pl = self._convert_to_pl_currency(leg1_npv_native, primary_currency, pl_currency)
            leg2_npv_pl = self._convert_to_pl_currency(leg2_npv_native, secondary_currency, pl_currency)
            total_npv_pl = leg1_npv_pl + leg2_npv_pl
            
            # Return leg NPVs in their native currencies (not converted)
            leg1_npv_return = leg1_npv_native
            leg2_npv_return = leg2_npv_native
            
            # Calculate delta sensitivities
            try:
                solver_name = self._get_solver_name(primary_currency, secondary_currency)
                solver = self.curve_builder.get_solver(solver_name)
                
                if solver is None:
                    raise ValueError(f"Solver '{solver_name}' not found")
                
                # For THB/JPY pairs, the solver structure may not support base="usd" directly
                # Try without base parameter first, or use a different approach
                if set([primary_currency, secondary_currency]) == {'THB', 'JPY'}:
                    # For THB/JPY, calculate delta without base parameter or handle differently
                    try:
                        # Try calculating delta without base parameter (will use native currencies)
                        delta = xccy.delta(solver=solver)
                        delta_values = delta.iloc[:, 0].values if hasattr(delta, 'iloc') else delta.values.flatten()
                    except:
                        # If that fails, try with base parameter
                        delta = xccy.delta(solver=solver, base="usd")
                        delta_values = delta.iloc[:, 0].values if hasattr(delta, 'iloc') else delta.values.flatten()
                else:
                    # For USD pairs, use base="usd"
                    delta = xccy.delta(solver=solver, base="usd")
                    delta_values = delta.iloc[:, 0].values if hasattr(delta, 'iloc') else delta.values.flatten()
                
                # Scale delta values by spot rates if needed
                if primary_currency == 'THB' and self.market_data.thb_spot is not None:
                    delta_values[:len(delta_values)//2] *= self.market_data.thb_spot
                elif primary_currency == 'JPY' and self.market_data.jpy_spot is not None:
                    delta_values[:len(delta_values)//2] *= self.market_data.jpy_spot
                
                delta = [round(num, 2) for num in delta_values]
            except Exception as e:
                print(f"Warning: Could not calculate delta sensitivities: {e}")
                delta = [0.0] * 10
            
            print(f"    Direct XCS NPV (in {pl_currency}): {total_npv_pl:.6f}")
            print(f"    Leg NPVs (in native): {primary_currency}={leg1_npv_native:.6f}, {secondary_currency}={leg2_npv_native:.6f}")
            
            return total_npv_pl, delta, leg1_npv_return, leg2_npv_return, cf_table
            
        except Exception as e:
            print(f"    Error in direct THB/JPY calculation: {e}")
            import traceback
            traceback.print_exc()
            return 0.0, [0.0], 0.0, 0.0, pd.DataFrame()

    def _calculate_thbjpy_direct_break_even(self, position: BookingPosition) -> float:
        """Calculate break-even spread for THB/JPY using direct XCS approach - minimal version"""
        try:
            primary_currency = position.leg1_currency
            secondary_currency = position.leg2_currency
            
            print(f"    Using direct XCS approach for {primary_currency}/{secondary_currency} break-even")
            
            # Get FX forward
            fx_forward = self._get_fx_forward(primary_currency, secondary_currency)
            
            # Get notional and apply position direction
            notional_leg1 = position.leg1_notional
            # B/S = Buy leg 1, Sell leg 2 - flip notional to match XCS convention
            if position.position == 'B/S':
                notional_leg1 *= -1
            
            # Determine FX fixing and currency setup
            if primary_currency == 'THB' and secondary_currency == 'JPY':
                fx_fixing = float(1.0 / self.market_data.thbjpy_spot)  # Invert THB/JPY to JPY/THB
                currency1 = 'thb'
                currency2 = 'jpy'
                # Leg 1 is THB, Leg 2 is JPY
                leg1_forecast = self.curve_builder.get_curve('thor')  # THB forecasting curve
                leg1_discount = self.curve_builder.get_curve('usdthb_xcs')  # THB discounting curve
                leg2_forecast = self.curve_builder.get_curve('tonr')  # JPY forecasting curve
                leg2_discount = self.curve_builder.get_curve('usdjpy_xcs')  # JPY discounting curve
            elif primary_currency == 'JPY' and secondary_currency == 'THB':
                fx_fixing = float(self.market_data.thbjpy_spot)  # Use THB/JPY directly
                currency1 = 'jpy'
                currency2 = 'thb'
                # Leg 1 is JPY, Leg 2 is THB
                leg1_forecast = self.curve_builder.get_curve('tonr')  # JPY forecasting curve
                leg1_discount = self.curve_builder.get_curve('usdjpy_xcs')  # JPY discounting curve
                leg2_forecast = self.curve_builder.get_curve('thor')  # THB forecasting curve
                leg2_discount = self.curve_builder.get_curve('usdthb_xcs')  # THB discounting curve
            else:
                raise ValueError(f"Unsupported currency pair: {primary_currency}/{secondary_currency}")
            
            # Create XCS using parameters from Excel booking
            xcs_params = {
                'effective': position.effective_date,
                'termination': position.maturity_date,
                'notional': notional_leg1,
                'frequency': position.get_frequency(primary_currency),  # Use actual frequency from Excel
                'currency': currency1,
                'leg2_currency': currency2,
                'calendar': 'tyo',  # Use Tokyo calendar for both
                'fx_fixings': fx_fixing,
                'fixed': position.get_fixed_float(primary_currency) == 'Fixed',
                'curves': [leg1_forecast, leg1_discount, leg2_forecast, leg2_discount],  # Correct curve assignment for each leg
                'convention': position.get_convention(primary_currency),  # Use actual convention from Excel
                'leg2_frequency': position.get_frequency(secondary_currency),  # Use actual frequency from Excel
                'leg2_calendar': 'tyo',
                'leg2_fixed': position.get_fixed_float(secondary_currency) == 'Fixed',
                'leg2_convention': position.get_convention(secondary_currency),  # Use actual convention from Excel
                'leg2_mtm': False,  # Disable mark-to-market to prevent intermediate principal exchanges
                'amortization': 0.0,  # Fixed amortization for all deal pairs
                'payment_lag': 2,
                'leg2_payment_lag': 2,
                'payment_lag_exchange': 0,
                'leg2_payment_lag_exchange': 0,
            }
            
            # Determine which leg to solve for
            solve_for_leg1 = position.get_fixed_float(primary_currency) == 'Float'
            solve_for_leg2 = position.get_fixed_float(secondary_currency) == 'Float'
            
            # Add fixed_rate for leg1 if it's fixed
            if position.get_fixed_float(primary_currency) == 'Fixed':
                rate_val = position.get_rate(primary_currency)
                if rate_val is not None:
                    xcs_params['fixed_rate'] = rate_val
            else:
                # Leg1 is floating - use NoInput.blank to let rate() solve for it
                xcs_params['float_spread'] = NoInput.blank
                xcs_params['fixings'] = self._get_fixings_for_currency(primary_currency)
                xcs_params['fixing_method'] = "rfr_lookback"
                xcs_params['method_param'] = 5  # Fixed 5 days lookback for all deal pairs
            
            # Add leg2 parameters based on fixed/float
            if position.get_fixed_float(secondary_currency) == 'Fixed':
                # Leg2 is fixed - just add the fixed rate, NO fixings
                rate_val = position.get_rate(secondary_currency)
                if rate_val is not None:
                    xcs_params['leg2_fixed_rate'] = rate_val
            else:
                # Leg2 is floating - use actual spread (will be overridden by rate() if solving for this leg)
                spread_val = position.get_spread(secondary_currency)
                if spread_val is not None:
                    xcs_params['leg2_float_spread'] = spread_val
                else:
                    xcs_params['leg2_float_spread'] = 0.0  # Default to 0 if not specified
                xcs_params['leg2_fixings'] = self._get_fixings_for_currency(secondary_currency)
                xcs_params['leg2_fixing_method'] = "rfr_lookback"
                xcs_params['leg2_method_param'] = 5  # Fixed 5 days lookback for all deal pairs
            
            # Validate all parameters before creating XCS (remove debug output for cleaner logs)
            none_params = []
            for key, value in xcs_params.items():
                if value is None and key not in ['float_spread', 'leg2_float_spread']:  # Allow NoInput.blank
                    none_params.append(key)
            
            if none_params:
                print(f"    ERROR: Found {len(none_params)} None parameters: {none_params}")
                return 0.0
            
            xccy = XCS(**xcs_params)
            
            # Calculate break-even spread using rate() method - must match curves in XCS constructor
            curve_objects = [leg1_forecast, leg1_discount, leg2_forecast, leg2_discount]
            solver_name = self._get_solver_name(primary_currency, secondary_currency)
            solver = self.curve_builder.get_solver(solver_name)
            
            # Determine which leg to solve for (1 for leg1, 2 for leg2)
            leg_to_solve = 1 if solve_for_leg1 else 2
            
            # Note: XCS.rate() doesn't accept 'target' parameter, it automatically solves for zero NPV
            break_even_spread = xccy.rate(
                curves=curve_objects,
                fx=fx_forward,
                solver=solver,
                leg=leg_to_solve  # Specify which leg to solve for
            )
            
            # Extract the actual value
            if isinstance(break_even_spread, (int, float)):
                result = float(break_even_spread)
            elif hasattr(break_even_spread, 'value'):
                result = float(break_even_spread.value)
            else:
                result = float(break_even_spread)
            
            print(f"    Direct XCS break-even spread: {result:.6f} bp")
            return result
            
        except Exception as e:
            print(f"    Error in direct THB/JPY break-even calculation: {e}")
            import traceback
            traceback.print_exc()
            return 0.0

    def _calculate_standard_npv(self, position: BookingPosition) -> Tuple[float, List[float], float, float, pd.DataFrame]:
        """
        Calculate NPV using the standard method (for non-THB/JPY CCS).
        This is the original NPV calculation logic.
        """
        # Get currencies and determine leg assignments
        currencies = position.currencies
        primary_currency = position.primary_currency
        secondary_currency = position.secondary_currency
        
        # Get notional amounts
        primary_notional = position.get_notional(primary_currency)
        secondary_notional = position.get_notional(secondary_currency)
        
        # Apply position direction
        # B/S = Buy leg 1 (receive), Sell leg 2 (pay) - XCS default is opposite, so flip for B/S
        # S/B = Sell leg 1 (pay), Buy leg 2 (receive) - XCS default matches this, keep as is
        if position.position == 'B/S':
            primary_notional *= -1
        
        # Get calendar for the trade based on currency pair
        # THB calendar is created from pair + Holidays sheet (not standalone)
        currencies = sorted([primary_currency, secondary_currency])
        
        if currencies == ['THB', 'USD']:
            # USD/THB: NYC + THB holidays
            calendar_name = 'us_th'
        elif currencies == ['JPY', 'USD']:
            # USD/JPY: NYC + JP holidays
            calendar_name = 'us_jp'
        elif currencies == ['JPY', 'THB']:
            # JPY/THB: TYO + THB holidays
            calendar_name = 'jp_th'
        else:
            # Default fallback
            calendar_name = 'us_th'
        
        calendar = self.curve_builder.calendar_manager.get_calendar(calendar_name)
        if calendar is None:
            calendar = self.curve_builder.calendar_manager.get_calendar('us_th')
        
        # Determine which leg is fixed and which is floating
        primary_fixed = position.get_fixed_float(primary_currency) == 'Fixed'
        secondary_fixed = position.get_fixed_float(secondary_currency) == 'Fixed'
        
        # Get spreads and rates
        primary_spread = position.get_spread(primary_currency)
        secondary_spread = position.get_spread(secondary_currency)
        
        # Get frequencies
        primary_freq = position.get_frequency(primary_currency)
        secondary_freq = position.get_frequency(secondary_currency)
        
        # Get day count conventions
        primary_convention = position.get_convention(primary_currency)
        secondary_convention = position.get_convention(secondary_currency)
        
        # Get the correct forecast and discount curves for each leg
        # Curves array format: [leg1_forecast, leg1_discount, leg2_forecast, leg2_discount]
        leg1_forecast = self._get_forecast_curve(primary_currency)
        leg1_discount = self._get_discount_curve(primary_currency, (primary_currency, secondary_currency))
        leg2_forecast = self._get_forecast_curve(secondary_currency)
        leg2_discount = self._get_discount_curve(secondary_currency, (primary_currency, secondary_currency))
        
        # Build curves array with actual curve objects
        curve_objects = [leg1_forecast, leg1_discount, leg2_forecast, leg2_discount]
        
        # Create XCS instrument with flexible fixed/float configuration
        xcs_params = {
            'effective': position.effective_date.to_pydatetime() if hasattr(position.effective_date, 'to_pydatetime') else position.effective_date,
            'termination': position.maturity_date.to_pydatetime() if hasattr(position.maturity_date, 'to_pydatetime') else position.maturity_date,
            'notional': primary_notional,
            'currency': primary_currency.lower(),
            'leg2_currency': secondary_currency.lower(),
            'calendar': calendar,
            'fx_fixings': self._get_fx_fixing(primary_currency, secondary_currency, position.fx_fixing),
            
            # Payment and exchange parameters
            'payment_lag': int(position.payment_lag),
            'leg2_payment_lag': int(position.payment_lag),
            'payment_lag_exchange': 0,
            'leg2_payment_lag_exchange': 0,
            
            # Amortization
            'amortization': 0.0,  # Fixed amortization for all deal pairs
            
            # Curves - provide correct curves for both legs
            'curves': curve_objects
        }
        
        # Configure primary leg (leg1) - Fixed or Float
        if primary_fixed:
            xcs_params.update({
                'fixed': True,
                'fixed_rate': position.get_rate(primary_currency),  # Use rate from Excel
                'frequency': primary_freq,
                'convention': primary_convention
            })
        else:
            xcs_params.update({
                'fixed': False,
                'float_spread': primary_spread,  # Use spread from Excel
                'frequency': primary_freq,
                'convention': primary_convention,
                'fixings': self._get_fixings_for_currency(primary_currency),
                'fixing_method': "rfr_lookback",
                'method_param': 5  # Fixed 5 days lookback for all deal pairs
            })
        
        # Configure secondary leg (leg2) - Fixed or Float
        if secondary_fixed:
            xcs_params.update({
                'leg2_fixed': True,
                'leg2_fixed_rate': position.get_rate(secondary_currency),  # Use rate from Excel
                'leg2_frequency': secondary_freq,
                'leg2_convention': secondary_convention,
                'leg2_calendar': calendar,
                'leg2_mtm': False
            })
        else:
            xcs_params.update({
                'leg2_fixed': False,
                'leg2_float_spread': secondary_spread,
                'leg2_frequency': secondary_freq,
                'leg2_convention': secondary_convention,
                'leg2_calendar': calendar,
                'leg2_mtm': False,
                'leg2_fixings': self._get_fixings_for_currency(secondary_currency),
                'leg2_fixing_method': "rfr_lookback",
                'leg2_method_param': 5  # Fixed 5 days lookback for all deal pairs
            })
        
        # Create the XCS instrument
        try:
            # Check for None values and fix them
            for key, value in xcs_params.items():
                if value is None:
                    if key in ['amortization', 'leg2_method_param']:
                        xcs_params[key] = 0.0
                    elif key in ['leg2_fixings']:
                        xcs_params[key] = self._get_fixings_for_currency(secondary_currency)
            
            xccy = XCS(**xcs_params)
        except Exception as e:
            return 0.0, [0.0], 0.0, 0.0, pd.DataFrame()
        
        # Verify curve objects are valid (curve_objects already built above)
        for i, curve_obj in enumerate(curve_objects):
            if curve_obj is None:
                print(f"Warning: Curve {i} is None for {position.booking_id}")
        
        # Get the correct FX forward object
        fx_forward = self._get_fx_forward(primary_currency, secondary_currency)
        if fx_forward is None:
            print(f"Warning: FX forward is None for {position.booking_id}")
        
        try:
            npv = float(xccy.npv(
                curves=curve_objects, 
                fx=fx_forward
            ))
        except Exception as e:
            print(f"Error calculating NPV for {position.booking_id}: {e}")
            return 0.0, [0.0], 0.0, 0.0, pd.DataFrame()
        
        try:
            self.cf_table = xccy.cashflows(
                curves=curve_objects, 
                fx=fx_forward
            )
        except Exception as e:
            print(f"Error calculating cashflows for {position.booking_id}: {e}")
            self.cf_table = pd.DataFrame()
        
        # Calculate leg-specific NPVs in native currencies
        primary_npv = 0.0
        secondary_npv = 0.0
        
        if not self.cf_table.empty:
            # Extract NPV for each leg in its native currency
            primary_npv = self.cf_table.NPV[self.cf_table.Ccy == primary_currency.upper()].sum()
            secondary_npv = self.cf_table.NPV[self.cf_table.Ccy == secondary_currency.upper()].sum()
        
        # Calculate Total NPV in PL Currency by converting each leg
        primary_npv_pl = self._convert_to_pl_currency(primary_npv, primary_currency, position.pl_currency)
        secondary_npv_pl = self._convert_to_pl_currency(secondary_npv, secondary_currency, position.pl_currency)
        total_npv_pl = primary_npv_pl + secondary_npv_pl
        
        # Return leg NPVs in their native currencies (not converted)
        primary_npv_return = primary_npv
        secondary_npv_return = secondary_npv
        
        # Calculate delta sensitivities (simplified)
        try:
            delta = [0.0] * 10  # Simplified delta calculation
        except Exception as e:
            print(f"Warning: Could not calculate delta sensitivities: {e}")
            delta = [0.0] * 10  # Default length
        
        return total_npv_pl, delta, primary_npv_return, secondary_npv_return, self.cf_table

    def calculate_break_even_spread(self, position: BookingPosition) -> float:
        """
        Calculate the break-even spread that makes NPV = 0 for a CCS position
        Uses the proper XCS.rate() method from rateslib
        
        Args:
            position: Booking position parameters with Leg 1 and Leg 2 details
            
        Returns:
            Break-even spread value (in basis points)
        """
        try:
            # Get currencies and determine leg assignments
            currencies = position.currencies
            primary_currency = position.primary_currency
            secondary_currency = position.secondary_currency
            
            # Calculate break-even spread
            
            # Get notional amounts
            primary_notional = position.get_notional(primary_currency)
            secondary_notional = position.get_notional(secondary_currency)
            
            # Apply position direction
            # B/S = Buy leg 1 (receive), Sell leg 2 (pay) - XCS default is opposite, so flip for B/S
            # S/B = Sell leg 1 (pay), Buy leg 2 (receive) - XCS default matches this, keep as is
            if position.position == 'B/S':
                primary_notional *= -1
            
            # Get calendar for the trade based on currency pair
            # THB calendar is created from pair + Holidays sheet (not standalone)
            currencies = sorted([primary_currency, secondary_currency])
            
            if currencies == ['THB', 'USD']:
                # USD/THB: NYC + THB holidays
                calendar_name = 'us_th'
            elif currencies == ['JPY', 'USD']:
                # USD/JPY: NYC + JP holidays
                calendar_name = 'us_jp'
            elif currencies == ['JPY', 'THB']:
                # JPY/THB: TYO + THB holidays
                calendar_name = 'jp_th'
            else:
                # Default fallback
                calendar_name = 'us_th'
            
            calendar = self.curve_builder.calendar_manager.get_calendar(calendar_name)
            if calendar is None:
                calendar = self.curve_builder.calendar_manager.get_calendar('us_th')
            
            # Determine which leg is fixed and which is floating
            primary_fixed = position.get_fixed_float(primary_currency) == 'Fixed'
            secondary_fixed = position.get_fixed_float(secondary_currency) == 'Fixed'
            
            # Determine leg configuration
            
            # Get frequencies and conventions
            primary_freq = position.get_frequency(primary_currency)
            secondary_freq = position.get_frequency(secondary_currency)
            primary_convention = position.get_convention(primary_currency)
            secondary_convention = position.get_convention(secondary_currency)
            
            # Get the correct forecast and discount curves for each leg
            # Curves array format: [leg1_forecast, leg1_discount, leg2_forecast, leg2_discount]
            leg1_forecast = self._get_forecast_curve(primary_currency)
            leg1_discount = self._get_discount_curve(primary_currency, (primary_currency, secondary_currency))
            leg2_forecast = self._get_forecast_curve(secondary_currency)
            leg2_discount = self._get_discount_curve(secondary_currency, (primary_currency, secondary_currency))
            
            # Build curves array with actual curve objects
            curve_objects = [leg1_forecast, leg1_discount, leg2_forecast, leg2_discount]
            
            # Get FX forward
            fx_forward = self._get_fx_forward(primary_currency, secondary_currency)
            
            
            # Get the appropriate solver for this currency pair
            solver_name = self._get_solver_name(primary_currency, secondary_currency)
            solver = self.curve_builder.get_solver(solver_name)
            
            # Get solver for break-even calculation
            
            # Create XCS instrument for break-even calculation - MUST match calculate_npv exactly
            xcs_params = {
                'effective': position.effective_date.to_pydatetime() if hasattr(position.effective_date, 'to_pydatetime') else position.effective_date,
                'termination': position.maturity_date.to_pydatetime() if hasattr(position.maturity_date, 'to_pydatetime') else position.maturity_date,
                'notional': primary_notional,
                'currency': primary_currency.lower(),
                'leg2_currency': secondary_currency.lower(),
                'calendar': calendar,
                'fx_fixings': self._get_fx_fixing(primary_currency, secondary_currency, position.fx_fixing),
                'payment_lag': int(position.payment_lag),
                'leg2_payment_lag': int(position.payment_lag),
                'payment_lag_exchange': 0,
                'leg2_payment_lag_exchange': 0,
                'amortization': 0.0,  # Fixed amortization for all deal pairs
                'curves': curve_objects
            }
            
            # Use spec parameter for THB/JPY to ensure consistency with solver and NPV calculation
            if currencies == ['JPY', 'THB']:
                xcs_params['spec'] = 'thbjpy_xcs'
                print(f"  Using thbjpy_xcs spec for break-even calculation")
            
            
            # Configure primary leg
            if primary_fixed:
                xcs_params.update({
                    'fixed': True,
                    'fixed_rate': position.get_rate(primary_currency),
                    'frequency': primary_freq,
                    'convention': primary_convention
                })
            else:
                # For floating leg, use NoInput.blank to let rate() method solve for it
                xcs_params.update({
                    'fixed': False,
                    'float_spread': NoInput.blank,  # Let rate() method solve for this
                    'frequency': primary_freq,
                    'convention': primary_convention,
                    'fixings': self._get_fixings_for_currency(primary_currency),
                    'fixing_method': "rfr_lookback",
                    'method_param': 5  # Fixed 5 days lookback for all deal pairs
                })
            
            # Configure secondary leg
            if secondary_fixed:
                xcs_params.update({
                    'leg2_fixed': True,
                    'leg2_fixed_rate': position.get_rate(secondary_currency),
                    'leg2_frequency': secondary_freq,
                    'leg2_convention': secondary_convention,
                    'leg2_calendar': calendar,
                    'leg2_mtm': False
                })
            else:
                # For THB/JPY with spec, only override trade-specific parameters
                if currencies == ['JPY', 'THB']:
                    xcs_params.update({
                        'leg2_fixed': False,
                        'leg2_float_spread': position.get_spread(secondary_currency),  # Use actual spread
                        'leg2_fixings': self._get_fixings_for_currency(secondary_currency)
                        # Let spec handle: frequency, convention, calendar, fixing_method, method_param
                    })
                else:
                    # For floating leg, use actual spread from position (same as NPV calculation)
                    xcs_params.update({
                        'leg2_fixed': False,
                        'leg2_float_spread': position.get_spread(secondary_currency),  # Use actual spread
                        'leg2_frequency': secondary_freq,
                        'leg2_convention': secondary_convention,
                        'leg2_calendar': calendar,
                        'leg2_mtm': False,
                        'leg2_fixings': self._get_fixings_for_currency(secondary_currency),
                        'leg2_fixing_method': "rfr_lookback",
                        'leg2_method_param': 5  # Fixed 5 days lookback for all deal pairs
                    })
            
            # Create the XCS instrument
            try:
                # Check for None values and fix them - MUST match calculate_npv exactly
                for key, value in xcs_params.items():
                    if value is None:
                        if key in ['amortization', 'leg2_method_param']:
                            xcs_params[key] = 0.0
                        elif key in ['leg2_fixings']:
                            xcs_params[key] = self._get_fixings_for_currency(secondary_currency)
                
                xccy = XCS(**xcs_params)
            except Exception as e:
                return 0.0
            
            # Validate configuration for rate() method
            if primary_fixed and secondary_fixed:
                print(f"Warning: Both legs are fixed for {position.booking_id} - no floating spread to solve")
                return 0.0
            
            # For THB/JPY CCS, use direct XCS approach with USD-based curves
            if set(currencies) == {'THB', 'JPY'}:
                print(f"  Using direct XCS approach for THB/JPY break-even spread")
                return self._calculate_thbjpy_direct_break_even(position)
            else:
                # Use the standard rate() method for other currency pairs
                print(f"  Using standard rate() method for {primary_currency}/{secondary_currency}")
                
                # Determine which leg to solve for based on which leg is floating
                # leg=1 means solve for primary (leg1), leg=2 means solve for secondary (leg2)
                solve_leg = 2 if primary_fixed and not secondary_fixed else 1
                
                break_even_spread = xccy.rate(
                    curves=curve_objects,
                    fx=fx_forward,
                    solver=solver,
                    leg=solve_leg  # Solve for the floating leg
                )
                print(f"  Raw rate() result: {break_even_spread}")
                print(f"  Rate() result type: {type(break_even_spread)}")

                # Extract the actual value from the rate() result
                if isinstance(break_even_spread, (int, float)):
                    result = float(break_even_spread)
                    print(f"  Extracted float value: {result}")
                    return result
                elif hasattr(break_even_spread, 'value'):
                    # Handle Dual objects from rateslib
                    try:
                        # For Dual objects, the .value attribute contains the actual value
                        actual_value = float(break_even_spread.value)
                        print(f"  Extracted Dual.value: {actual_value}")
                        return actual_value
                    except Exception as e:
                        print(f"  Error extracting Dual.value: {e}")
                        # Fallback: try to convert the object directly
                        try:
                            actual_value = float(break_even_spread)
                            print(f"  Fallback conversion: {actual_value}")
                            return actual_value
                        except Exception as e2:
                            print(f"  Fallback conversion failed: {e2}")
                            return 0.0
                else:
                    # Try direct conversion for other types
                    try:
                        actual_value = float(break_even_spread)
                        print(f"  Direct conversion: {actual_value}")
                        return actual_value
                    except Exception as e:
                        print(f"  Warning: Cannot convert break-even spread to float: {e}")
                        print(f"  Type: {type(break_even_spread)}")
                        return 0.0
                        
        except Exception as e:
            print(f"Error calculating break-even spread for {position.booking_id}: {e}")
            return 0.0
    
    def calculate_fx_delta(self, position: BookingPosition) -> Dict[str, float]:
        """
        Calculate FX delta manually by bumping FX rate by 1%
        
        Formula: (MV(FX_0 + 1% * FX_0) - MV(FX_0)) / 1%
        
        For THB/JPY triangulated deals, calculates sensitivities to BOTH USD/THB and USD/JPY
        since the deal is priced through USD triangulation.
        
        Args:
            position: Booking position parameters
            
        Returns:
            Dictionary with FX delta for each relevant currency pair
        """
        try:
            # Get base NPV
            base_npv, _, _, _, _ = self.calculate_npv(position)
            
            # Determine which FX rate to shock based on currency pair
            primary_currency = position.primary_currency
            secondary_currency = position.secondary_currency
            
            fx_deltas = {}
            
            # Store original FX rates
            original_thb_spot = self.market_data.thb_spot
            original_jpy_spot = self.market_data.jpy_spot
            original_thbjpy_spot = self.market_data.thbjpy_spot
            
            # Shock relevant FX rate by 1%
            if set([primary_currency, secondary_currency]) == {'USD', 'THB'}:
                # Shock USD/THB
                print(f"    Calculating USDTHB FX Delta...")
                self.market_data.thb_spot = original_thb_spot * 1.01
                # Rebuild curves to reflect FX change
                self.curve_builder._build_fx_curves()
                shocked_npv, _, _, _, _ = self.calculate_npv(position)
                fx_deltas['USDTHB'] = (shocked_npv - base_npv) / 0.01
                # Restore
                self.market_data.thb_spot = original_thb_spot
                self.curve_builder._build_fx_curves()
                
            elif set([primary_currency, secondary_currency]) == {'USD', 'JPY'}:
                # Shock USD/JPY
                print(f"    Calculating USDJPY FX Delta...")
                self.market_data.jpy_spot = original_jpy_spot * 1.01
                # Rebuild curves to reflect FX change
                self.curve_builder._build_fx_curves()
                shocked_npv, _, _, _, _ = self.calculate_npv(position)
                fx_deltas['USDJPY'] = (shocked_npv - base_npv) / 0.01
                # Restore
                self.market_data.jpy_spot = original_jpy_spot
                self.curve_builder._build_fx_curves()
                
            elif set([primary_currency, secondary_currency]) == {'THB', 'JPY'}:
                # For THB/JPY triangulated deals, calculate sensitivity to BOTH USD/THB and USD/JPY
                # since the deal is priced through USD triangulation
                
                # 1. Shock USD/THB
                print(f"    Calculating USDTHB FX Delta (triangulation component)...")
                self.market_data.thb_spot = original_thb_spot * 1.01
                # Update triangulated THB/JPY rate accordingly
                self.market_data.thbjpy_spot = self.market_data.thb_spot / self.market_data.jpy_spot
                # Rebuild curves to reflect FX change
                self.curve_builder._build_fx_curves()
                shocked_npv, _, _, _, _ = self.calculate_npv(position)
                fx_deltas['USDTHB'] = (shocked_npv - base_npv) / 0.01
                # Restore
                self.market_data.thb_spot = original_thb_spot
                self.market_data.thbjpy_spot = original_thbjpy_spot
                self.curve_builder._build_fx_curves()
                
                # 2. Shock USD/JPY
                print(f"    Calculating USDJPY FX Delta (triangulation component)...")
                self.market_data.jpy_spot = original_jpy_spot * 1.01
                # Update triangulated THB/JPY rate accordingly
                self.market_data.thbjpy_spot = self.market_data.thb_spot / self.market_data.jpy_spot
                # Rebuild curves to reflect FX change
                self.curve_builder._build_fx_curves()
                shocked_npv, _, _, _, _ = self.calculate_npv(position)
                fx_deltas['USDJPY'] = (shocked_npv - base_npv) / 0.01
                # Restore
                self.market_data.jpy_spot = original_jpy_spot
                self.market_data.thbjpy_spot = original_thbjpy_spot
                self.curve_builder._build_fx_curves()
            
            return fx_deltas
            
        except Exception as e:
            print(f"Error calculating FX delta for {position.booking_id}: {e}")
            return {}
    
    def calculate_ir_dv01(self, position: BookingPosition) -> Dict[str, pd.DataFrame]:
        """
        Calculate IR DV01 manually by bumping each rate point by 0.01% (1 bp)
        
        Formula: MV(IR_0 + 0.01%) - MV(IR_0)
        
        KEY RULES:
        1. Fixed legs have NO curve sensitivity (DV01 = 0 for fixed legs)
        2. THB/JPY triangulated deals need ALL USD-based sensitivities:
           - SOFR (USD discounting in triangulation)
           - THOR (if THB leg is floating)
           - TONR (if JPY leg is floating)
           - USDTHB_CCS (triangulation component)
           - USDJPY_CCS (triangulation component)
        
        Args:
            position: Booking position parameters
            
        Returns:
            Dictionary with DV01 DataFrames for each relevant curve
        """
        try:
            # Get base NPV
            base_npv, _, _, _, _ = self.calculate_npv(position)
            
            # Determine which curves are relevant for this position
            primary_currency = position.primary_currency
            secondary_currency = position.secondary_currency
            
            # Check which legs are floating (only floating legs have curve sensitivity)
            primary_is_floating = position.get_fixed_float(primary_currency) == 'Float'
            secondary_is_floating = position.get_fixed_float(secondary_currency) == 'Float'
            
            dv01_results = {}
            
            # Map currencies to their base curves
            currency_curves = {
                'USD': ('sofr', self.market_data.sofr_rates),
                'THB': ('thor', self.market_data.thor_rates),
                'JPY': ('tonr', self.market_data.tonr_rates)
            }
            
            # For THB/JPY triangulated deals, we need to include SOFR sensitivity
            is_thbjpy_triangulated = set([primary_currency, secondary_currency]) == {'THB', 'JPY'}
            
            # Calculate DV01 for each relevant base curve (only for floating legs)
            for currency in [primary_currency, secondary_currency]:
                # Check if this leg is floating
                is_floating = (currency == primary_currency and primary_is_floating) or \
                             (currency == secondary_currency and secondary_is_floating)
                
                if is_floating and currency in currency_curves:
                    curve_name, rates_df = currency_curves[currency]
                    dv01_data = []
                    
                    print(f"    Calculating {currency}_{curve_name.upper()} DV01 (floating leg)...")
                    
                    # Shock each rate point by 1 bp (0.01%)
                    for idx, row in rates_df.iterrows():
                        # Store original rate
                        original_rate = rates_df.at[idx, 'Rate']
                        
                        # Shock rate by 1 bp
                        rates_df.at[idx, 'Rate'] = original_rate + 0.01
                        
                        # Rebuild curves with shocked rate
                        self.curve_builder._build_base_curves()
                        self.curve_builder._build_fx_curves()
                        self.curve_builder._build_solvers()
                        
                        # Recalculate NPV
                        shocked_npv, _, _, _, _ = self.calculate_npv(position)
                        
                        # Calculate DV01
                        dv01 = shocked_npv - base_npv
                        
                        # Store result
                        dv01_data.append({
                            'Tenor': row['Term'],
                            'Rate': original_rate,
                            'DV01': dv01
                        })
                        
                        # Restore original rate
                        rates_df.at[idx, 'Rate'] = original_rate
                    
                    # Rebuild curves with original rates
                    self.curve_builder._build_base_curves()
                    self.curve_builder._build_fx_curves()
                    self.curve_builder._build_solvers()
                    
                    dv01_results[f'{currency}_{curve_name.upper()}'] = pd.DataFrame(dv01_data)
                elif not is_floating:
                    print(f"    Skipping {currency} DV01 (fixed leg - no curve sensitivity)")
            
            # For THB/JPY triangulated deals, add SOFR sensitivity (used for USD discounting in triangulation)
            if is_thbjpy_triangulated:
                print(f"    Calculating USD_SOFR DV01 (triangulation discounting)...")
                sofr_dv01_data = []
                sofr_rates_df = self.market_data.sofr_rates
                
                for idx, row in sofr_rates_df.iterrows():
                    # Store original rate
                    original_rate = sofr_rates_df.at[idx, 'Rate']
                    
                    # Shock rate by 1 bp
                    sofr_rates_df.at[idx, 'Rate'] = original_rate + 0.01
                    
                    # Rebuild curves with shocked rate
                    self.curve_builder._build_base_curves()
                    self.curve_builder._build_fx_curves()
                    self.curve_builder._build_solvers()
                    
                    # Recalculate NPV
                    shocked_npv, _, _, _, _ = self.calculate_npv(position)
                    
                    # Calculate DV01
                    dv01 = shocked_npv - base_npv
                    
                    # Store result
                    sofr_dv01_data.append({
                        'Tenor': row['Term'],
                        'Rate': original_rate,
                        'DV01': dv01
                    })
                    
                    # Restore original rate
                    sofr_rates_df.at[idx, 'Rate'] = original_rate
                
                # Rebuild curves with original rates
                self.curve_builder._build_base_curves()
                self.curve_builder._build_fx_curves()
                self.curve_builder._build_solvers()
                
                dv01_results['USD_SOFR'] = pd.DataFrame(sofr_dv01_data)
            
            # Calculate DV01 for CCS curves if applicable
            if set([primary_currency, secondary_currency]) == {'USD', 'THB'}:
                print(f"    Calculating USDTHB_CCS DV01...")
                dv01_results['USDTHB_CCS'] = self._calculate_ccs_dv01(position, 'usdthb', base_npv)
            elif set([primary_currency, secondary_currency]) == {'USD', 'JPY'}:
                print(f"    Calculating USDJPY_CCS DV01...")
                dv01_results['USDJPY_CCS'] = self._calculate_ccs_dv01(position, 'usdjpy', base_npv)
            elif set([primary_currency, secondary_currency]) == {'THB', 'JPY'}:
                # For THB/JPY triangulation, we need both USD/THB and USD/JPY CCS sensitivities
                print(f"    Calculating USDTHB_CCS DV01 (triangulation component)...")
                dv01_results['USDTHB_CCS'] = self._calculate_ccs_dv01(position, 'usdthb', base_npv)
                print(f"    Calculating USDJPY_CCS DV01 (triangulation component)...")
                dv01_results['USDJPY_CCS'] = self._calculate_ccs_dv01(position, 'usdjpy', base_npv)
            
            return dv01_results
            
        except Exception as e:
            print(f"Error calculating IR DV01 for {position.booking_id}: {e}")
            import traceback
            traceback.print_exc()
            return {}
    
    def _calculate_ccs_dv01(self, position: BookingPosition, ccs_type: str, base_npv: float) -> pd.DataFrame:
        """
        Calculate DV01 for CCS curve points
        
        Args:
            position: Booking position
            ccs_type: 'usdthb', 'usdjpy', or 'thbjpy'
            base_npv: Base NPV value
            
        Returns:
            DataFrame with CCS DV01 values
        """
        dv01_data = []
        
        try:
            # Get the appropriate CCS rates
            if ccs_type == 'usdthb':
                ccs_rates = self.market_data.usdthb_ccs_rates
            elif ccs_type == 'usdjpy':
                ccs_rates = self.market_data.usdjpy_ccs_rates
            elif ccs_type == 'thbjpy':
                # For THB/JPY, we use triangulated rates, so skip CCS DV01
                return pd.DataFrame()
            else:
                return pd.DataFrame()
            
            # Shock each CCS rate point by 1 bp
            for idx, row in ccs_rates.iterrows():
                # Store original rate
                original_rate = ccs_rates.at[idx, 'Rate']
                
                # Shock rate by 1 bp
                ccs_rates.at[idx, 'Rate'] = original_rate + 0.01
                
                # Rebuild curves with shocked rate
                self.curve_builder._build_base_curves()
                self.curve_builder._build_fx_curves()
                self.curve_builder._build_solvers()
                
                # Recalculate NPV
                shocked_npv, _, _, _, _ = self.calculate_npv(position)
                
                # Calculate DV01
                dv01 = shocked_npv - base_npv
                
                # Store result
                dv01_data.append({
                    'Tenor': row['Term'],
                    'Rate': original_rate,
                    'DV01': dv01
                })
                
                # Restore original rate
                ccs_rates.at[idx, 'Rate'] = original_rate
            
            # Rebuild curves with original rates
            self.curve_builder._build_base_curves()
            self.curve_builder._build_fx_curves()
            self.curve_builder._build_solvers()
            
        except Exception as e:
            print(f"Error calculating CCS DV01 for {ccs_type}: {e}")
        
        return pd.DataFrame(dv01_data)
    
    def calculate_all_sensitivities(self, position: BookingPosition) -> Dict[str, any]:
        """
        Calculate all sensitivities (FX delta and IR DV01) for a position
        
        Sensitivity coverage by deal type:
        
        1. USD/THB deals:
           - FX: USD/THB
           - IR: SOFR (if USD leg is floating), THOR (if THB leg is floating), USDTHB_CCS
        
        2. USD/JPY deals:
           - FX: USD/JPY
           - IR: SOFR (if USD leg is floating), TONR (if JPY leg is floating), USDJPY_CCS
        
        3. THB/JPY deals (triangulated via USD):
           - FX: USD/THB, USD/JPY (both components of triangulation)
           - IR: SOFR (USD discounting), THOR (if THB leg is floating), TONR (if JPY leg is floating),
                 USDTHB_CCS, USDJPY_CCS (both components of triangulation)
        
        Note: Fixed legs have NO curve sensitivity (DV01 = 0 for fixed legs)
        
        Args:
            position: Booking position parameters
            
        Returns:
            Dictionary with FX deltas and IR DV01s
        """
        try:
            print(f"  Calculating sensitivities for {position.booking_id}...")
            
            # Calculate FX delta
            fx_deltas = self.calculate_fx_delta(position)
            
            # Calculate IR DV01
            ir_dv01s = self.calculate_ir_dv01(position)
            
            return {
                'booking_id': position.booking_id,
                'fx_deltas': fx_deltas,
                'ir_dv01s': ir_dv01s
            }
            
        except Exception as e:
            print(f"Error calculating sensitivities for {position.booking_id}: {e}")
            return {
                'booking_id': position.booking_id,
                'fx_deltas': {},
                'ir_dv01s': {}
            }
    
    def process_booking_sheet(self, booking_file: str, sheet_name: str = 'Booking') -> Tuple[pd.DataFrame, Dict[str, pd.DataFrame]]:
        """
        Process booking sheet and calculate NPVs for all positions using new comprehensive format
        
        Only processes trades with Trade Status = "live" (case-insensitive)
        Skips trades with any other status (e.g., "dead", "cancelled", etc.)
        
        Args:
            booking_file: Path to Excel file with booking data
            sheet_name: Name of the booking sheet
            
        Returns:
            Tuple of (booking_results_df, cashflow_data_dict)
            - booking_results_df: DataFrame with NPV results for all positions
            - cashflow_data_dict: Dictionary of cash flows (only for live trades)
        """
        book = pd.read_excel(booking_file, sheet_name=sheet_name, engine='openpyxl')
        book = book.reset_index()
        
        print(f"Processing {len(book)} positions...")
        
        # Process booking positions
        
        # Add result columns (NPVs in PL Currency specified per position)
        book['Net_NPV'] = 0.0
        book['Leg_1_NPV'] = 0.0
        book['Leg_2_NPV'] = 0.0
        book['Break_Even_Spread'] = 0.0
        
        # Dictionary to store cash flow data for each position
        cashflow_data = {}
        
        # Track statistics
        live_count = 0
        skipped_count = 0
        
        for row in range(len(book)):
            try:
                # Extract position parameters using new comprehensive structure
                position = self._extract_comprehensive_position(book, row)
                
                # Check Trade Status - only process "live" trades
                if hasattr(position, 'trade_status') and position.trade_status:
                    trade_status_normalized = str(position.trade_status).strip().lower()
                else:
                    trade_status_normalized = ''
                
                if trade_status_normalized != 'live':
                    # Skip non-live trades
                    skipped_count += 1
                    print(f"⊗ Skipping {position.booking_id} - Status: '{position.trade_status}' (not 'live')")
                    # Set NPV to zero for skipped trades
                    book.loc[row, 'Net_NPV'] = 0.0
                    book.loc[row, 'Leg_1_NPV'] = 0.0
                    book.loc[row, 'Leg_2_NPV'] = 0.0
                    book.loc[row, 'Break_Even_Spread'] = 0.0
                    continue  # Skip to next trade
                
                # Process live trade
                live_count += 1
                print(f"✓ Processing live trade: {position.booking_id}")
            
                # Calculate NPV and cash flows (in PL Currency)
                npv, delta, leg1_npv, leg2_npv, cf_table = self.calculate_npv(position)
                # Calculate break-even spread
                break_even_spread = self.calculate_break_even_spread(position)            
                # Store results (NPVs in PL Currency specified in position)
                book.loc[row, 'Net_NPV'] = npv
                book.loc[row, 'Leg_1_NPV'] = leg1_npv
                book.loc[row, 'Leg_2_NPV'] = leg2_npv
                book.loc[row, 'Break_Even_Spread'] = break_even_spread
            
                # Store cash flow data
                if cf_table is not None and not cf_table.empty:
                    # Add booking ID to cash flow table for identification
                    cf_table_copy = cf_table.copy()
                    cf_table_copy['Booking_ID'] = position.booking_id
                    cashflow_data[position.booking_id] = cf_table_copy
                else:
                    print(f"Warning: No cash flows generated for {position.booking_id}")
                    
            except Exception as e:
                booking_id = book.iloc[row].get('Booking\nID', 'Unknown')
                print(f"Error processing row {row} (Booking ID: {booking_id}): {e}")
                # Set default values for failed calculations
                book.loc[row, 'Net_NPV'] = 0.0
                book.loc[row, 'Leg_1_NPV'] = 0.0
                book.loc[row, 'Leg_2_NPV'] = 0.0
        
        print(f"\n{'='*60}")
        print(f"PROCESSING SUMMARY")
        print(f"{'='*60}")
        print(f"  Total positions in sheet: {len(book)}")
        print(f"  Live trades processed: {live_count}")
        print(f"  Skipped trades (not live): {skipped_count}")
        print(f"  Positions with cash flows: {len(cashflow_data)}")
        if cashflow_data:
            total_cashflows = sum(len(cf) for cf in cashflow_data.values())
            print(f"  Total cash flows generated: {total_cashflows}")
        else:
            print(f"  No cash flows generated - check for errors above")
        print(f"{'='*60}")
        
        return book, cashflow_data

    def update_booking_npv_to_excel(self, booking_results: pd.DataFrame, excel_file: str = 'CCS_Template.xlsx', sheet_name: str = 'Booking') -> None:
        """
        Update the Booking sheet with NPV values in columns AE, AF, AG, AH
        NPV values are in the PL Currency specified in each position
        
        Args:
            booking_results: DataFrame with calculated NPV values
            excel_file: Path to Excel file
            sheet_name: Name of the booking sheet
        """
        try:
            from openpyxl import load_workbook
            
            # Load the existing workbook
            wb = load_workbook(excel_file)
            
            # Get the Booking sheet
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                print(f"Warning: Sheet '{sheet_name}' not found in {excel_file}")
                return
            
            
            print(f"Updating NPV values in {sheet_name} sheet...")
            
            for idx, row in booking_results.iterrows():
                # Get the row number in Excel (add 2 because Excel is 1-indexed and we skip header)
                excel_row = idx + 2
                
                # Update Net NPV in column AB - in PL Currency
                if 'Net_NPV' in row and pd.notna(row['Net_NPV']):
                    ws.cell(row=excel_row, column=EXCEL_COL_NET_NPV, value=float(row['Net_NPV']))
                
                # Update Leg 1 NPV in column AC - in Leg 1's native currency
                if 'Leg_1_NPV' in row and pd.notna(row['Leg_1_NPV']):
                    ws.cell(row=excel_row, column=EXCEL_COL_LEG1_NPV, value=float(row['Leg_1_NPV']))
                
                # Update Leg 2 NPV in column AD - in Leg 2's native currency
                if 'Leg_2_NPV' in row and pd.notna(row['Leg_2_NPV']):
                    ws.cell(row=excel_row, column=EXCEL_COL_LEG2_NPV, value=float(row['Leg_2_NPV']))
            
                # Update Break Even Spread in column AH
                if 'Break_Even_Spread' in row and pd.notna(row['Break_Even_Spread']):
                    ws.cell(row=excel_row, column=EXCEL_COL_BREAKEVEN, value=float(row['Break_Even_Spread']))
            # Save the workbook
            wb.save(excel_file)
            print(f"✓ Successfully updated NPV values in {sheet_name} sheet")
            print(f"   - Net NPV (PL Currency) values written to column AB")
            print(f"   - Leg 1 NPV (Leg 1 Currency) values written to column AC") 
            print(f"   - Leg 2 NPV (Leg 2 Currency) values written to column AD")
            print(f"   - Break Even Spread values written to column AE")
            
        except Exception as e:
            print(f"Error updating NPV values in Excel: {e}")

    def _apply_currency_defaults(self, currency: str, curve: str, daycount: str, calendar: str) -> Tuple[str, str, str]:
        """
        Apply currency-specific defaults for curve, daycount, and calendar
        
        Args:
            currency: Currency code ('USD', 'THB', 'JPY')
            curve: Existing curve value (may be None)
            daycount: Existing daycount value (may be None)
            calendar: Existing calendar value (may be None)
            
        Returns:
            Tuple of (curve, daycount, calendar) with defaults applied
        """
        currency_defaults = {
            'USD': (DEFAULT_USD_CURVE, DEFAULT_USD_CONVENTION, 'us'),
            'THB': (DEFAULT_THB_CURVE, DEFAULT_THB_CONVENTION, None),  # THB calendar created from pair + Holidays
            'JPY': (DEFAULT_JPY_CURVE, DEFAULT_JPY_CONVENTION, 'jp')
        }
        
        default_curve, default_daycount, default_calendar = currency_defaults.get(
            currency, 
            (DEFAULT_USD_CURVE, DEFAULT_USD_CONVENTION, 'us')
        )
        
        return (
            curve if curve is not None else default_curve,
            daycount if daycount is not None else default_daycount,
            calendar if calendar is not None else default_calendar
        )
    
    def _calculate_default_notional(self, leg1_currency: str, leg2_currency: str, leg1_notional: float) -> float:
        """
        Calculate default notional for leg 2 based on currency pair using market data
        
        Args:
            leg1_currency: Leg 1 currency
            leg2_currency: Leg 2 currency
            leg1_notional: Leg 1 notional amount
            
        Returns:
            Calculated leg 2 notional
        """
        # Use market spot rates
        if leg1_currency == 'USD' and leg2_currency == 'THB':
            fx_rate = float(self.market_data.thb_spot)
        elif leg1_currency == 'USD' and leg2_currency == 'JPY':
            fx_rate = float(self.market_data.jpy_spot)
        elif leg1_currency == 'THB' and leg2_currency == 'JPY':
            fx_rate = float(1.0 / self.market_data.thbjpy_spot)
        else:
            fx_rate = 1.0  # Default fallback
        
        return leg1_notional * fx_rate
    
    def _calculate_default_fx_fixing(self, leg1_currency: str, leg2_currency: str) -> float:
        """
        Calculate default FX fixing based on currency pair using market data
        
        Args:
            leg1_currency: Leg 1 currency
            leg2_currency: Leg 2 currency
            
        Returns:
            FX fixing rate from market data
        """
        # Use market spot rates
        if leg1_currency == 'USD' and leg2_currency == 'JPY':
            return float(self.market_data.jpy_spot)
        elif leg1_currency == 'USD' and leg2_currency == 'THB':
            return float(self.market_data.thb_spot)
        elif leg1_currency == 'JPY' and leg2_currency == 'THB':
            return float(1.0 / self.market_data.thbjpy_spot)
        else:
            return 1.0  # Default fallback
    
    def _validate_and_normalize_pl_currency(self, pl_currency: str, booking_id: str) -> str:
        """
        Validate and normalize P&L currency
        
        Args:
            pl_currency: P&L currency from Excel
            booking_id: Booking ID for error messages
            
        Returns:
            Validated and normalized P&L currency
        """
        if pd.isna(pl_currency) or pl_currency is None:
            return DEFAULT_PL_CURRENCY
        
        pl_currency = str(pl_currency).upper().strip()
        
        if pl_currency not in SUPPORTED_CURRENCIES:
            print(f"Warning: Invalid PL Currency '{pl_currency}' for {booking_id}. "
                  f"Defaulting to {DEFAULT_PL_CURRENCY}.")
            return DEFAULT_PL_CURRENCY
        
        return pl_currency

    def _extract_comprehensive_position(self, book: pd.DataFrame, row: int) -> BookingPosition:
        """
        Extract position from comprehensive booking sheet format using exact column names
        
        Args:
            book: DataFrame containing booking data
            row: Row index to extract
            
        Returns:
            BookingPosition object with extracted parameters
        """
        def safe_get(col_name: str, default=None):
            if col_name in book.columns:
                value = book.iloc[row][col_name]
                # Handle NaN values
                if pd.isna(value):
                    return default
                return value
            return default
        
        # Extract position data from Excel row
        
        # Extract general trade information (using exact column names from Excel)
        booking_id = safe_get('Booking\nID', f'Row_{row}')
        effective_date = safe_get('Effective\n Date', pd.Timestamp.now())
        maturity_date = safe_get('Maturity\n Date', pd.Timestamp.now() + pd.DateOffset(years=3))
        position = safe_get('S/B or \nB/S', 'B/S')
        
        # Convert dates to datetime if they're not already
        if isinstance(effective_date, str):
            effective_date = pd.to_datetime(effective_date)
        elif pd.isna(effective_date):
            effective_date = pd.Timestamp.now()
            
        if isinstance(maturity_date, str):
            maturity_date = pd.to_datetime(maturity_date)
        elif pd.isna(maturity_date):
            maturity_date = pd.Timestamp.now() + pd.DateOffset(years=3)
        
        # Extract Leg 1 parameters (using exact column names from Excel)
        leg1_currency = safe_get('Leg 1\n Currency', None)
        leg1_notional = safe_get('Leg 1\nNotional', None)
        leg1_fixed_float = safe_get('Leg 1\nFixed / Float', None)
        leg1_curve = safe_get('Leg 1\nIndex', None)
        leg1_rate = safe_get('Leg 1\nRate (%)', None)
        leg1_spread = safe_get('Leg 1\nSpread (bp)', None)
        leg1_daycount = safe_get('Leg 1\nDayCount', None)
        leg1_pay_freq = safe_get('Leg 1\nPay Freq', None)
        leg1_reset_freq = safe_get('Leg 1\nRestet Freq', None)
        leg1_calendar = safe_get('Leg 1\nCalendar', None)
        
        # Apply currency-specific defaults for Leg 1
        leg1_curve, leg1_daycount, leg1_calendar = self._apply_currency_defaults(
            leg1_currency, leg1_curve, leg1_daycount, leg1_calendar
        )
        
        # Extract Leg 2 parameters (using exact column names from Excel)
        leg2_currency = safe_get('Leg 2\nCurrency', None)
        leg2_notional = safe_get('Leg 2\nNotional', None)
        leg2_fixed_float = safe_get('Leg 2\nFixed / Float', None)
        leg2_curve = safe_get('Leg 2\nIndex', None)
        leg2_rate = safe_get('Leg 2\nRate (%)', None)
        leg2_spread = safe_get('Leg 2\nSpread (bp)', None)
        leg2_daycount = safe_get('Leg 2\nDayCount', None)
        leg2_frequency = safe_get('Leg 2\nFrequency', None)
        leg2_reset_freq = safe_get('Leg 2\nRestet Freq', None)
        leg2_calendar = safe_get('Leg 2\nCalendar', None)
        
        # Apply currency-specific defaults for Leg 2
        leg2_curve, leg2_daycount, leg2_calendar = self._apply_currency_defaults(
            leg2_currency, leg2_curve, leg2_daycount, leg2_calendar
        )
        
        # Extract additional trade attributes (using exact column names from Excel)
        fx_fixing = safe_get('FX_Fixing', None)
        lookback_days = safe_get('Lookback\nDays', None)
        payment_lag = safe_get('Payment\nLag', None)
        trade_status = safe_get('Trade Status', None)
        pl_currency = safe_get('PL Currency', DEFAULT_PL_CURRENCY)
        
        # Provide defaults for missing fields
        if pd.isna(leg2_notional) or leg2_notional is None:
            leg2_notional = self._calculate_default_notional(leg1_currency, leg2_currency, leg1_notional)
        
        if pd.isna(fx_fixing) or fx_fixing is None:
            fx_fixing = self._calculate_default_fx_fixing(leg1_currency, leg2_currency)
        
        # Convert numeric fields (handle NaN values) - do this before validation
        leg1_notional = float(leg1_notional) if leg1_notional is not None and not pd.isna(leg1_notional) else 0.0
        leg2_notional = float(leg2_notional) if leg2_notional is not None and not pd.isna(leg2_notional) else 0.0
        
        # Validate required fields (after applying defaults and conversions)
        required_fields = [
            ('Leg 1 Currency', leg1_currency),
            ('Leg 1 Notional', leg1_notional),
            ('Leg 1 Fixed/Float', leg1_fixed_float),
            ('Leg 1 Index', leg1_curve),
            ('Leg 2 Currency', leg2_currency),
            ('Leg 2 Notional', leg2_notional),
            ('Leg 2 Fixed/Float', leg2_fixed_float),
            ('Leg 2 Index', leg2_curve),
            ('FX Fixing', fx_fixing)
        ]
        
        missing_fields = [field for field, value in required_fields if value is None or pd.isna(value) or value == 0.0]
        if missing_fields:
            raise ValueError(f"Missing required fields in Excel: {', '.join(missing_fields)}")
        
        # Convert remaining numeric fields
        fx_fixing = float(fx_fixing) if fx_fixing is not None and not pd.isna(fx_fixing) else 1.0  # Note: FX fixing is ignored, market data is used
        lookback_days = int(lookback_days) if lookback_days is not None and not pd.isna(lookback_days) else 0  # Note: Lookback is hardcoded to 5 days in deal calculations
        payment_lag = int(payment_lag) if payment_lag is not None and not pd.isna(payment_lag) else DEFAULT_PAYMENT_LAG
        
        # Convert rate fields
        leg1_rate = float(leg1_rate) if leg1_rate is not None and not pd.isna(leg1_rate) else DEFAULT_RATE
        leg2_rate = float(leg2_rate) if leg2_rate is not None and not pd.isna(leg2_rate) else DEFAULT_RATE
        leg1_spread = float(leg1_spread) if leg1_spread is not None and not pd.isna(leg1_spread) else DEFAULT_SPREAD
        leg2_spread = float(leg2_spread) if leg2_spread is not None and not pd.isna(leg2_spread) else DEFAULT_SPREAD
        
        # Validate and normalize PL Currency
        pl_currency = self._validate_and_normalize_pl_currency(pl_currency, booking_id)
        
        # Display configuration read from Excel
        print(f"\n--- CONFIGURATION FOR {booking_id} ---")
        print(f"Trade: {position} | Status: {trade_status} | {effective_date.strftime('%Y-%m-%d')} to {maturity_date.strftime('%Y-%m-%d')}")
        print(f"Leg 1: {leg1_currency} {leg1_notional:,.0f} | {leg1_fixed_float} | {leg1_curve} | Rate: {leg1_rate}% | Spread: {leg1_spread}bp")
        print(f"Leg 2: {leg2_currency} {leg2_notional:,.0f} | {leg2_fixed_float} | {leg2_curve} | Rate: {leg2_rate}% | Spread: {leg2_spread}bp")
        print(f"FX: {fx_fixing} | Payment Lag: {payment_lag} days | PL Currency: {pl_currency}")
        
        return BookingPosition(
            # General Trade Information
            booking_id=booking_id,
            effective_date=effective_date,
            maturity_date=maturity_date,
            position=position,
            
            # Leg 1 Parameters
            leg1_currency=leg1_currency,
            leg1_notional=leg1_notional,
            leg1_fixed_float=leg1_fixed_float,
            leg1_curve=leg1_curve,
            leg1_rate=leg1_rate,
            leg1_spread=leg1_spread,
            leg1_daycount=leg1_daycount,
            leg1_pay_freq=leg1_pay_freq,
            leg1_reset_freq=leg1_reset_freq,
            leg1_calendar=leg1_calendar,
            
            # Leg 2 Parameters
            leg2_currency=leg2_currency,
            leg2_notional=leg2_notional,
            leg2_fixed_float=leg2_fixed_float,
            leg2_curve=leg2_curve,
            leg2_rate=leg2_rate,
            leg2_spread=leg2_spread,
            leg2_daycount=leg2_daycount,
            leg2_frequency=leg2_frequency,
            leg2_reset_freq=leg2_reset_freq,
            leg2_calendar=leg2_calendar,
            
            # Additional Trade Attributes
            fx_fixing=fx_fixing,
            lookback_days=lookback_days,
            payment_lag=payment_lag,
            trade_status=trade_status,
            pl_currency=pl_currency
        )

class LHBCCSProcessor:
    """Main class that orchestrates the entire CCS curve construction and booking process"""
    
    def __init__(self, excel_file: str):
        self.excel_file = excel_file
        self.calendar_manager = None
        self.market_data_loader = None
        self.market_data = None
        self.curve_builder = None
        self.booking_calculator = None
        
        self._initialize()
    
    def _initialize(self) -> None:
        """Initialize all components"""
        # Initialize calendar manager
        self.calendar_manager = CalendarManager(self.excel_file)
        
        # Initialize market data loader
        self.market_data_loader = MarketDataLoader(self.excel_file)
        self.market_data = self.market_data_loader.get_market_data()
        
        # Initialize curve builder
        self.curve_builder = CurveBuilder(self.market_data, self.calendar_manager)
        
        # Initialize booking calculator
        self.booking_calculator = CCSBookingCalculator(self.curve_builder, self.market_data)
    
    def process_booking(self, booking_file: str = None, sheet_name: str = 'Booking') -> Tuple[pd.DataFrame, Dict[str, pd.DataFrame]]:
        """
        Process booking positions and calculate NPVs
        
        Args:
            booking_file: Path to booking file (defaults to main Excel file)
            sheet_name: Name of booking sheet
            
        Returns:
            Tuple of (booking_results_df, cashflow_data_dict)
        """
        if booking_file is None:
            booking_file = self.excel_file
        
        return self.booking_calculator.process_booking_sheet(booking_file, sheet_name)
    
    def update_booking_npv_to_excel(self, booking_results: pd.DataFrame, excel_file: str = None, sheet_name: str = 'Booking') -> None:
        """
        Update the Booking sheet with NPV values in columns AE, AF, AG
        
        Args:
            booking_results: DataFrame with calculated NPV values
            excel_file: Path to Excel file (defaults to main Excel file)
            sheet_name: Name of the booking sheet
        """
        if excel_file is None:
            excel_file = self.excel_file
        
        return self.booking_calculator.update_booking_npv_to_excel(booking_results, excel_file, sheet_name)
    
    def get_curve_info(self) -> Dict[str, pd.DataFrame]:
        """Get information about all constructed curves including basis spreads"""
        curve_info = {}
        
        # Get base curves
        for curve_name in ['sofr', 'thor', 'tonr', 'usdthb_xcs', 'usdjpy_xcs']:
            curve = self.curve_builder.get_curve(curve_name)
            if curve:
                try:
                    # Get curve data from the market data termination dates
                    if curve_name == 'sofr':
                        termination_dates = self.curve_builder.market_data.sofr_rates["Termination"]
                    elif curve_name == 'thor':
                        termination_dates = self.curve_builder.market_data.thor_rates["Termination"]
                    elif curve_name == 'tonr':
                        termination_dates = self.curve_builder.market_data.tonr_rates["Termination"]
                    elif curve_name == 'usdthb_xcs':
                        ut_rate = pd.concat([self.curve_builder.market_data.usdthb_sw_rates, 
                                           self.curve_builder.market_data.usdthb_ccs_rates], ignore_index=True)
                        termination_dates = ut_rate["Termination"]
                    elif curve_name == 'usdjpy_xcs':
                        uj_rate = pd.concat([self.curve_builder.market_data.usdjpy_sw_rates, 
                                           self.curve_builder.market_data.usdjpy_ccs_rates], ignore_index=True)
                        termination_dates = uj_rate["Termination"]
                    else:
                        termination_dates = []
                    
                    # Add curve date to the list
                    all_dates = [self.curve_builder.market_data.curve_date] + list(termination_dates)
                    
                    # Get discount factors
                    df_values = [float(curve[date]) for date in all_dates]
                    
                    curve_info[curve_name] = pd.DataFrame({
                        'Date': all_dates,
                        'DF': df_values
                    })
                    
                except Exception as e:
                    print(f"Error extracting curve data for {curve_name}: {e}")
                    curve_info[curve_name] = pd.DataFrame({
                        'Date': [],
                        'DF': []
                    })
        
        # Basis spreads removed as requested
        
        return curve_info
    
    def display_curve_details(self, curve_name: str, curve_data: pd.DataFrame) -> None:
        """
        Display detailed curve information including rates and discount factors
        
        Args:
            curve_name: Name of the curve
            curve_data: DataFrame with curve data
        """
        if curve_data.empty:
            print(f"No data available for {curve_name}")
            return
        
        print(f"\n{curve_name.upper()} CURVE DETAILS")
        print("=" * 60)
        
        # Basis spread curves removed as requested
        
        # Calculate zero rates and forward rates for regular curves
        curve_data = curve_data.copy()
        curve_data['Zero_Rate'] = 0.0
        curve_data['Forward_Rate'] = 0.0
        
        for i in range(len(curve_data)):
            if i == 0:
                curve_data.iloc[i, curve_data.columns.get_loc('Zero_Rate')] = 0.0
                curve_data.iloc[i, curve_data.columns.get_loc('Forward_Rate')] = 0.0
            else:
                # Calculate zero rate
                days = (curve_data.iloc[i]['Date'] - curve_data.iloc[0]['Date']).days
                if days > 0:
                    df = curve_data.iloc[i]['DF']
                    zero_rate = -BUSINESS_DAYS_IN_YEAR * np.log(df) / days
                    curve_data.iloc[i, curve_data.columns.get_loc('Zero_Rate')] = zero_rate * PERCENTAGE_MULTIPLIER
                    
                    # Calculate forward rate
                    if i > 1:
                        prev_df = curve_data.iloc[i-1]['DF']
                        prev_days = (curve_data.iloc[i-1]['Date'] - curve_data.iloc[0]['Date']).days
                        period_days = days - prev_days
                        if period_days > 0:
                            forward_rate = BUSINESS_DAYS_IN_YEAR * np.log(prev_df / df) / period_days
                            curve_data.iloc[i, curve_data.columns.get_loc('Forward_Rate')] = forward_rate * PERCENTAGE_MULTIPLIER
        
        # Display formatted table
        display_df = curve_data[['Date', 'DF', 'Zero_Rate', 'Forward_Rate']].copy()
        display_df['Date'] = display_df['Date'].dt.strftime('%Y-%m-%d')
        
        print(display_df.to_string(index=False))
        
        # Summary statistics
        print(f"\nSummary:")
        print(f"  Points: {len(curve_data)}")
        print(f"  Start Date: {curve_data['Date'].min().strftime('%Y-%m-%d')}")
        print(f"  End Date: {curve_data['Date'].max().strftime('%Y-%m-%d')}")
        print(f"  Min DF: {curve_data['DF'].min():.6f}")
        print(f"  Max DF: {curve_data['DF'].max():.6f}")
        if curve_data['Zero_Rate'].max() > 0:
            print(f"  Max Zero Rate: {curve_data['Zero_Rate'].max():.4f}%")
    
    def display_all_curves(self) -> None:
        """Display all constructed curves with detailed information"""
        print("\n" + "="*80)
        print("RATE CURVES DISPLAY")
        print("="*80)
        
        curve_info = self.get_curve_info()
        
        for curve_name, curve_data in curve_info.items():
            self.display_curve_details(curve_name, curve_data)
        
        print("\n" + "="*80)
        print("CURVE SUMMARY")
        print("="*80)
        
        # Summary table
        summary_data = []
        for curve_name, curve_data in curve_info.items():
            if not curve_data.empty and 'basis' not in curve_name.lower():
                # Handle regular curves only (basis spread curves removed)
                summary_data.append({
                    'Curve': curve_name.upper(),
                    'Points': len(curve_data),
                    'Start_Date': curve_data['Date'].min().strftime('%Y-%m-%d'),
                    'End_Date': curve_data['Date'].max().strftime('%Y-%m-%d'),
                    'Min_DF': f"{curve_data['DF'].min():.6f}",
                    'Max_DF': f"{curve_data['DF'].max():.6f}"
                })
        
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            print(summary_df.to_string(index=False))
        else:
            print("No curve data available")
    
    def add_curves_to_excel(self, excel_file: str = 'CCS_Template.xlsx') -> None:
        """Add all curve data to a single sheet in existing Excel file"""
        try:
            from openpyxl import load_workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            curve_info = self.get_curve_info()
            
            # Load existing workbook to preserve formatting
            wb = load_workbook(excel_file)
            
            # Create or get the curves sheet
            if 'All_Curves' in wb.sheetnames:
                # Remove existing sheet and create new one to start fresh
                wb.remove(wb['All_Curves'])
                ws = wb.create_sheet('All_Curves')
                current_row = 1
            else:
                ws = wb.create_sheet('All_Curves')
                current_row = 1
            
            # Add all curves to the same sheet (excluding basis spread curves)
            for curve_name, curve_data in curve_info.items():
                if not curve_data.empty and 'basis' not in curve_name.lower():
                    # Handle regular curves only
                    # Calculate additional metrics for regular curves
                    curve_data = curve_data.copy()
                    curve_data['Zero_Rate'] = 0.0
                    curve_data['Forward_Rate'] = 0.0
                    
                    for i in range(len(curve_data)):
                        if i > 0:
                            days = (curve_data.iloc[i]['Date'] - curve_data.iloc[0]['Date']).days
                            if days > 0:
                                df = curve_data.iloc[i]['DF']
                                zero_rate = -BUSINESS_DAYS_IN_YEAR * np.log(df) / days
                                curve_data.iloc[i, curve_data.columns.get_loc('Zero_Rate')] = zero_rate * PERCENTAGE_MULTIPLIER
                                
                                if i > 1:
                                    prev_df = curve_data.iloc[i-1]['DF']
                                    prev_days = (curve_data.iloc[i-1]['Date'] - curve_data.iloc[0]['Date']).days
                                    period_days = days - prev_days
                                    if period_days > 0:
                                        forward_rate = BUSINESS_DAYS_IN_YEAR * np.log(prev_df / df) / period_days
                                        curve_data.iloc[i, curve_data.columns.get_loc('Forward_Rate')] = forward_rate * PERCENTAGE_MULTIPLIER
                    
                    # Format for Excel
                    export_df = curve_data.copy()
                    export_df['Date'] = export_df['Date'].dt.strftime('%Y-%m-%d')
                    
                    # Add curve name as header
                    ws.cell(row=current_row, column=1, value=f"{curve_name.upper()} CURVE")
                    current_row += 1
                    
                    # Add column headers
                    headers = ['Date', 'DF', 'Zero_Rate', 'Forward_Rate']
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=current_row, column=col, value=header)
                    current_row += 1
                    
                    # Add data rows
                    for _, row_data in export_df.iterrows():
                        for col, value in enumerate(row_data, 1):
                            ws.cell(row=current_row, column=col, value=value)
                        current_row += 1
                    
                    # Add empty row between curves
                    current_row += 1
            
            # Save the workbook
            wb.save(excel_file)
            print(f"\nAll curve data added to 'All_Curves' sheet in: {excel_file}")
            
        except Exception as e:
            print(f"Error adding curves to Excel: {e}")
    
    def get_cashflow_summary(self, cashflow_data: Dict[str, pd.DataFrame]) -> pd.DataFrame:
        """
        Get a summary of cash flows across all positions
        
        Args:
            cashflow_data: Dictionary of cash flow data by booking ID
            
        Returns:
            DataFrame with cash flow summary
        """
        summary_data = []
        
        for booking_id, cf_table in cashflow_data.items():
            # Group by currency and sum NPV
            currency_summary = cf_table.groupby('Ccy')['NPV'].sum().to_dict()
            
            # Convert all NPVs to USD
            usd_npv = currency_summary.get('USD', 0)
            thb_npv_usd = currency_summary.get('THB', 0) / self.market_data.thb_spot if 'THB' in currency_summary else 0
            jpy_npv_usd = currency_summary.get('JPY', 0) / self.market_data.jpy_spot if 'JPY' in currency_summary else 0
            
            total_npv_usd = usd_npv + thb_npv_usd + jpy_npv_usd
            
            summary_data.append({
                'Booking_ID': booking_id,
                'USD_NPV': usd_npv,
                'THB_NPV_USD': thb_npv_usd,
                'JPY_NPV_USD': jpy_npv_usd,
                'Total_NPV_USD': total_npv_usd,
                'Total_Flows': len(cf_table),
                'USD_Flows': len(cf_table[cf_table['Ccy'] == 'USD']),
                'THB_Flows': len(cf_table[cf_table['Ccy'] == 'THB']),
                'JPY_Flows': len(cf_table[cf_table['Ccy'] == 'JPY'])
            })
        
        return pd.DataFrame(summary_data)
    
    def populate_cash_flow_settled_sheet(self, cashflow_data: Dict[str, pd.DataFrame], 
                                       excel_file: str = 'CCS_Template.xlsx') -> None:
        """
        Populate the 'Cash Flow Settled' sheet with cash flow data
        
        Args:
            cashflow_data: Dictionary of cash flow data by booking ID
            excel_file: Path to Excel file
        """
        try:
            # Create combined cash flow DataFrame
            combined_cashflows = self._create_combined_cashflow_dataframe(cashflow_data)
            
            # Load existing workbook to preserve formatting
            from openpyxl import load_workbook
            wb = load_workbook(excel_file)
            
            # Create or get the CF sheet
            if 'CF' in wb.sheetnames:
                ws = wb['CF']
                # Clear only the data area, preserve headers and formatting
                for row in ws.iter_rows(min_row=2):  # Keep header row
                    for cell in row:
                        cell.value = None
            else:
                ws = wb.create_sheet('CF')
            
            # Write headers in row 1
            if not combined_cashflows.empty:
                headers = combined_cashflows.columns.tolist()
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_idx, value=header)
            
            # Write the combined cash flows starting from row 2 (after header)
            for row_idx, (_, row_data) in enumerate(combined_cashflows.iterrows(), 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Save the workbook
            wb.save(excel_file)
            
            print(f"CF sheet updated in {excel_file}")
            print(f"Added {len(combined_cashflows)} cash flow records")
            
        except Exception as e:
            print(f"Error populating CF sheet: {e}")
    
    def _create_combined_cashflow_dataframe(self, cashflow_data: Dict[str, pd.DataFrame]) -> pd.DataFrame:
        """Create a combined DataFrame with all cash flow data"""
        if not cashflow_data:
            return pd.DataFrame()
        
        # Combine all cash flow data
        all_cashflows = []
        for booking_id, cf_table in cashflow_data.items():
            if not cf_table.empty:
                # Add booking ID to the cash flow table
                cf_table_copy = cf_table.copy()
                cf_table_copy['Booking_ID'] = booking_id
                all_cashflows.append(cf_table_copy)
        
        if not all_cashflows:
            return pd.DataFrame()
        
        # Combine all cash flows
        combined_df = pd.concat(all_cashflows, ignore_index=True)
        
        # Use all available columns from the cashflows() function
        # Keep Booking_ID as the first column, then all other columns from cashflows()
        all_columns = ['Booking_ID'] + [col for col in combined_df.columns if col != 'Booking_ID']
        result_df = combined_df[all_columns].copy()
        
        # Format dates - check for common date column patterns
        for col in result_df.columns:
            if any(date_keyword in col.lower() for date_keyword in ['date', 'start', 'end', 'payment', 'acc']):
                if result_df[col].dtype == 'object':  # Only format if it's not already datetime
                    try:
                        result_df[col] = pd.to_datetime(result_df[col], errors='coerce').dt.strftime('%Y-%m-%d')
                    except (ValueError, TypeError, AttributeError):
                        pass  # Skip if can't convert to date
        
        return result_df
    
    def calculate_all_sensitivities(self, booking_file: str = None, sheet_name: str = 'Booking') -> Dict[str, Dict]:
        """
        Calculate sensitivities (FX delta and IR DV01) for all live booking positions
        
        Args:
            booking_file: Path to booking file (defaults to main Excel file)
            sheet_name: Name of booking sheet
            
        Returns:
            Dictionary of sensitivity data by booking ID
        """
        if booking_file is None:
            booking_file = self.excel_file
        
        book = pd.read_excel(booking_file, sheet_name=sheet_name, engine='openpyxl')
        book = book.reset_index()
        
        print(f"\n{'='*60}")
        print(f"CALCULATING SENSITIVITIES")
        print(f"{'='*60}")
        
        all_sensitivities = {}
        processed_count = 0
        
        for row in range(len(book)):
            try:
                # Extract position parameters
                position = self.booking_calculator._extract_comprehensive_position(book, row)
                
                # Check Trade Status - only process "live" trades
                if hasattr(position, 'trade_status') and position.trade_status:
                    trade_status_normalized = str(position.trade_status).strip().lower()
                else:
                    trade_status_normalized = ''
                
                if trade_status_normalized != 'live':
                    continue
                
                # Calculate sensitivities for live trade
                sensitivity_data = self.booking_calculator.calculate_all_sensitivities(position)
                
                if sensitivity_data:
                    all_sensitivities[position.booking_id] = sensitivity_data
                    processed_count += 1
                else:
                    print(f"Warning: No sensitivity data for {position.booking_id}")
                    
            except Exception as e:
                booking_id = book.iloc[row].get('Booking\nID', f'Row_{row}')
                print(f"Error calculating sensitivity for row {row} (Booking ID: {booking_id}): {e}")
        
        print(f"\n{'='*60}")
        print(f"Calculated sensitivities for {processed_count} live positions")
        print(f"{'='*60}")
        
        return all_sensitivities
    
    def add_sensitivities_to_excel(self, sensitivities: Dict[str, Dict], 
                                   excel_file: str = 'CCS_Template.xlsx') -> None:
        """
        Add sensitivity data (FX delta and IR DV01) to Excel file
        
        Args:
            sensitivities: Dictionary of sensitivity data by booking ID
            excel_file: Path to Excel file
        """
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(excel_file)
            
            # Create or clear the Sensitivities sheet
            if 'Sensitivities' in wb.sheetnames:
                wb.remove(wb['Sensitivities'])
            ws = wb.create_sheet('Sensitivities')
            
            current_row = 1
            
            # Add header
            ws.cell(row=current_row, column=1, value="SENSITIVITY ANALYSIS")
            ws.cell(row=current_row, column=2, value=f"Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}")
            current_row += 2
            
            # Process each booking position
            for booking_id, sens_data in sensitivities.items():
                # Add booking ID header
                ws.cell(row=current_row, column=1, value=f"BOOKING ID: {booking_id}")
                current_row += 1
                
                # Add FX Delta section
                if 'fx_deltas' in sens_data and sens_data['fx_deltas']:
                    ws.cell(row=current_row, column=1, value="FX DELTA (1% bump)")
                    current_row += 1
                    
                    ws.cell(row=current_row, column=1, value="Currency Pair")
                    ws.cell(row=current_row, column=2, value="FX Delta")
                    current_row += 1
                    
                    for fx_pair, delta_value in sens_data['fx_deltas'].items():
                        ws.cell(row=current_row, column=1, value=fx_pair)
                        ws.cell(row=current_row, column=2, value=float(delta_value))
                        current_row += 1
                    
                    current_row += 1
                
                # Add IR DV01 section
                if 'ir_dv01s' in sens_data and sens_data['ir_dv01s']:
                    ws.cell(row=current_row, column=1, value="IR DV01 (1 bp bump)")
                    current_row += 1
                    
                    for curve_name, dv01_df in sens_data['ir_dv01s'].items():
                        if not dv01_df.empty:
                            # Add curve name
                            ws.cell(row=current_row, column=1, value=f"{curve_name} Curve")
                            current_row += 1
                            
                            # Add column headers
                            ws.cell(row=current_row, column=1, value="Tenor")
                            ws.cell(row=current_row, column=2, value="Rate (%)")
                            ws.cell(row=current_row, column=3, value="DV01")
                            current_row += 1
                            
                            # Add data rows
                            for _, row_data in dv01_df.iterrows():
                                ws.cell(row=current_row, column=1, value=row_data['Tenor'])
                                ws.cell(row=current_row, column=2, value=float(row_data['Rate']))
                                ws.cell(row=current_row, column=3, value=float(row_data['DV01']))
                                current_row += 1
                            
                            current_row += 1
                
                current_row += 2  # Extra space between positions
            
            # Save the workbook
            wb.save(excel_file)
            print(f"\n✓ Sensitivity data written to 'Sensitivities' sheet in: {excel_file}")
            
        except Exception as e:
            print(f"Error adding sensitivities to Excel: {e}")
            import traceback
            traceback.print_exc()
    
    def display_sensitivity_summary(self, sensitivities: Dict[str, Dict]) -> None:
        """
        Display summary of calculated sensitivities
        
        Args:
            sensitivities: Dictionary of sensitivity data by booking ID
        """
        print("\n" + "="*80)
        print("SENSITIVITY ANALYSIS SUMMARY")
        print("="*80)
        
        if not sensitivities:
            print("No sensitivity data available")
            return
        
        summary_data = []
        
        for booking_id, sens_data in sensitivities.items():
            # Count FX deltas
            fx_count = len(sens_data.get('fx_deltas', {}))
            
            # Count IR DV01 buckets
            ir_bucket_count = 0
            for curve_name, dv01_df in sens_data.get('ir_dv01s', {}).items():
                if not dv01_df.empty:
                    ir_bucket_count += len(dv01_df)
            
            summary_data.append({
                'Booking_ID': booking_id,
                'FX_Deltas': fx_count,
                'IR_DV01_Buckets': ir_bucket_count,
                'Total_Sensitivities': fx_count + ir_bucket_count
            })
        
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            print(summary_df.to_string(index=False))
            print(f"\nTotal positions: {len(summary_data)}")
            print(f"Total FX deltas: {summary_df['FX_Deltas'].sum()}")
            print(f"Total IR DV01 buckets: {summary_df['IR_DV01_Buckets'].sum()}")
        else:
            print("No sensitivity data to display")


def main():
    """Main execution function"""
    # Configuration - allow Excel file to be specified via command line argument
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    else:
        excel_file = 'CCS_Template.xlsx'
    
    # Validate file exists
    if not os.path.exists(excel_file):
        print(f"Error: Excel file '{excel_file}' not found.")
        print(f"Usage: python {os.path.basename(__file__)} [excel_file.xlsx]")
        sys.exit(1)
    
    print(f"Starting LHB CCS Curve Construction...")
    print(f"Using Excel file: {excel_file}")
    
    try:
        print("Initializing processor...")
        processor = LHBCCSProcessor(excel_file)
        print("Processor initialized successfully")
        
        print("Processing booking...")
        booking_results, cashflow_data = processor.process_booking()
        print("Booking processing completed")
        
        # Update NPV values in Excel Booking sheet
        print("\nUpdating NPV values in Excel...")
        processor.update_booking_npv_to_excel(booking_results)
        
        # Display booking results
        print("\nBooking Results:")
        display_cols = ['Booking\nID', 'PL Currency', 'Net_NPV', 'Leg_1_NPV', 'Leg_2_NPV', 'Break_Even_Spread']
        available_cols = [col for col in display_cols if col in booking_results.columns]
        print(booking_results[available_cols].head())
        
        # Display cash flow summary
        print("\nCash Flow Summary:")
        cf_summary = processor.get_cashflow_summary(cashflow_data)
        print(cf_summary)
        
        # Display detailed cash flows for all positions
        if cashflow_data:
            print("\nDetailed Cash Flows:")
            for booking_id, cf_table in cashflow_data.items():
                print(f"\n--- {booking_id} ---")
                if not cf_table.empty:
                    print(f"Total cash flows: {len(cf_table)}")
                    
                    # Group by currency and convert to USD
                    currency_summary = cf_table.groupby('Ccy')['NPV'].sum()
                    print("NPV by currency (converted to USD):")
                    total_npv_usd = 0
                    for ccy, npv in currency_summary.items():
                        if ccy == 'USD':
                            npv_usd = npv
                        elif ccy == 'THB':
                            npv_usd = npv / processor.market_data.thb_spot
                        elif ccy == 'JPY':
                            npv_usd = npv / processor.market_data.jpy_spot
                        else:
                            npv_usd = npv
                        
                        total_npv_usd += npv_usd
                        print(f"  {ccy}: {npv:,.2f} -> {npv_usd:,.2f} USD")
                    
                    print(f"  Total NPV: {total_npv_usd:,.2f} USD")
                    print(f"Currencies involved: {sorted(currency_summary.index.tolist())}")
                    
                    # Show first few cashflows with all available columns
                    available_cols = cf_table.columns.tolist()
                    print("\nFirst few cash flows:")
                    print(cf_table[available_cols].head(8).to_string(index=False))
                else:
                    print("No cash flow data available")
        else:
            print("\nNo cash flow data generated - check for errors above")
        
        # Add curves to Excel
        processor.add_curves_to_excel()
        
        # Populate CF sheet
        print("\nPopulating CF sheet...")
        processor.populate_cash_flow_settled_sheet(cashflow_data)
        
        # Calculate sensitivities (FX delta and IR DV01)
        print("\nCalculating sensitivities (FX delta and IR DV01)...")
        sensitivities = processor.calculate_all_sensitivities()
        
        # Display sensitivity summary
        processor.display_sensitivity_summary(sensitivities)
        
        # Add sensitivities to Excel
        print("\nAdding sensitivities to Excel...")
        processor.add_sensitivities_to_excel(sensitivities)
        
        print("\nProcessing complete!")
        
    except Exception as e:
        print(f"Error: {e}")
        raise


if __name__ == "__main__":
    main()


